from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from pathlib import Path
import os, logging, json, csv, io

from core import db, master_db, client, get_tenant_db, get_user_db, JWT_SECRET, JWT_ALGORITHM, ROOT_DIR
from core.auth import (
    security, hash_password, verify_password, generate_id, generate_order_no,
    create_token, get_current_user, log_audit
)
from core.models import *

router = APIRouter()
UPLOAD_DIR = Path(ROOT_DIR) / "uploads"

# ==================== Reports Routes ====================

@router.get("/reports/sales-summary")
async def get_sales_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    store_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if store_id:
        query["store_id"] = store_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    sales_orders = await udb.sales_orders.find(query, {"_id": 0}).to_list(10000)
    online_orders = await udb.online_orders.find(
        {"order_status": "completed"} if not query else {**query, "order_status": "completed"},
        {"_id": 0}
    ).to_list(10000)
    
    total_sales = sum(o["total_amount"] for o in sales_orders)
    total_online = sum(o["total_amount"] for o in online_orders)
    
    by_method = {}
    for o in sales_orders:
        method = o.get("payment_method", "other")
        if method not in by_method:
            by_method[method] = {"count": 0, "amount": 0.0}
        by_method[method]["count"] += 1
        by_method[method]["amount"] += o.get("total_amount", 0)
    
    return {
        "total_sales": total_sales,
        "total_online_sales": total_online,
        "total_combined": total_sales + total_online,
        "sales_count": len(sales_orders),
        "online_count": len(online_orders),
        "by_payment_method": by_method
    }

@router.get("/reports/inventory-summary")
async def get_inventory_summary(
    warehouse_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    
    inventory = await udb.inventory.find(query, {"_id": 0}).to_list(10000)
    
    total_items = len(inventory)
    total_quantity = sum(i["quantity"] for i in inventory)
    total_value = 0.0
    low_stock_items = []
    
    for inv in inventory:
        product = await udb.products.find_one({"id": inv["product_id"]}, {"_id": 0})
        if product:
            total_value += inv["quantity"] * product.get("cost_price", 0)
            if inv["quantity"] <= product.get("min_stock", 0):
                low_stock_items.append({
                    "product": product,
                    "quantity": inv["quantity"],
                    "min_stock": product.get("min_stock", 0)
                })
    
    return {
        "total_items": total_items,
        "total_quantity": total_quantity,
        "total_value": total_value,
        "low_stock_count": len(low_stock_items),
        "low_stock_items": low_stock_items[:10]
    }

@router.get("/reports/top-products")
async def get_top_products(
    limit: int = 10,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    sales_orders = await udb.sales_orders.find(query, {"_id": 0}).to_list(10000)
    
    product_sales = {}
    for order in sales_orders:
        for item in order.get("items", []):
            pid = item["product_id"]
            if pid not in product_sales:
                product_sales[pid] = {"quantity": 0, "amount": 0}
            product_sales[pid]["quantity"] += item["quantity"]
            product_sales[pid]["amount"] += item["amount"]
    
    # Sort by amount
    sorted_products = sorted(product_sales.items(), key=lambda x: x[1]["amount"], reverse=True)[:limit]
    
    result = []
    for pid, stats in sorted_products:
        product = await udb.products.find_one({"id": pid}, {"_id": 0})
        if product:
            result.append({
                "product": product,
                "quantity": stats["quantity"],
                "amount": stats["amount"]
            })
    
    return result

# ==================== Dashboard Routes ====================

@router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    
    # Today's sales
    today_sales = await udb.sales_orders.find({"created_at": {"$gte": today}}, {"_id": 0}).to_list(1000)
    today_online = await udb.online_orders.find({"created_at": {"$gte": today}}, {"_id": 0}).to_list(1000)
    
    # Counts
    products_count = await udb.products.count_documents({})
    customers_count = await udb.customers.count_documents({})
    stores_count = await udb.stores.count_documents({})
    pending_orders = await udb.online_orders.count_documents({"order_status": "pending"})
    
    return {
        "today_sales_amount": sum(o["total_amount"] for o in today_sales),
        "today_sales_count": len(today_sales),
        "today_online_amount": sum(o["total_amount"] for o in today_online),
        "today_online_count": len(today_online),
        "products_count": products_count,
        "customers_count": customers_count,
        "stores_count": stores_count,
        "pending_online_orders": pending_orders
    }

# ==================== Reports: Top/Bottom Products ====================

@router.get("/reports/bestsellers")
async def get_bestsellers(days: int = 30, limit: int = 20, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    orders = await udb.sales_orders.find({"created_at": {"$gte": since}}, {"_id": 0}).to_list(10000)
    
    product_sales = {}
    for o in orders:
        for item in o.get("items", []):
            pid = item.get("product_id", "")
            if pid not in product_sales:
                product_sales[pid] = {"quantity": 0, "amount": 0, "name": item.get("product_name", pid)}
            product_sales[pid]["quantity"] += item.get("quantity", 0)
            product_sales[pid]["amount"] += item.get("amount", 0)
    
    sorted_products = sorted(product_sales.items(), key=lambda x: x[1]["amount"], reverse=True)
    bestsellers = [{"product_id": k, **v} for k, v in sorted_products[:limit]]
    slowsellers = [{"product_id": k, **v} for k, v in sorted_products[-limit:] if v["amount"] > 0]
    
    return {"bestsellers": bestsellers, "slowsellers": slowsellers}

# ==================== Profit Analysis ====================
@router.get("/reports/profit-analysis")
async def get_profit_analysis(
    start_date: Optional[str] = None, end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if start_date: query["created_at"] = {"$gte": start_date}
    if end_date: query.setdefault("created_at", {})["$lte"] = end_date
    orders = await udb.sales_orders.find(query, {"_id": 0}).to_list(10000)
    product_profit = {}
    for order in orders:
        for item in order.get("items", []):
            pid = item["product_id"]
            if pid not in product_profit:
                product_profit[pid] = {"revenue": 0, "cost": 0, "quantity": 0}
            product_profit[pid]["revenue"] += item.get("amount", 0)
            product_profit[pid]["quantity"] += item.get("quantity", 0)
    products = await udb.products.find({}, {"_id": 0}).to_list(10000)
    prod_map = {p["id"]: p for p in products}
    result = []
    total_revenue = 0
    total_cost = 0
    for pid, data in product_profit.items():
        product = prod_map.get(pid, {})
        cost = product.get("cost_price", 0) * data["quantity"]
        profit = data["revenue"] - cost
        margin = (profit / data["revenue"] * 100) if data["revenue"] > 0 else 0
        total_revenue += data["revenue"]
        total_cost += cost
        result.append({
            "product_id": pid, "product_name": product.get("name", "?"),
            "product_code": product.get("code", ""),
            "quantity": data["quantity"], "revenue": round(data["revenue"], 2),
            "cost": round(cost, 2), "profit": round(profit, 2), "margin": round(margin, 1)
        })
    result.sort(key=lambda x: x["profit"], reverse=True)
    total_profit = total_revenue - total_cost
    overall_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    return {
        "items": result, "total_revenue": round(total_revenue, 2),
        "total_cost": round(total_cost, 2), "total_profit": round(total_profit, 2),
        "overall_margin": round(overall_margin, 1)
    }

# ==================== Tax Report ====================
@router.get("/reports/tax")
async def get_tax_report(start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    query = {"status": "completed"}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date
    orders = await udb.sales_orders.find(query, {"_id": 0}).to_list(10000)
    tax_summary = {}
    total_sales = 0
    total_tax = 0
    total_base = 0
    for order in orders:
        total_sales += order.get("total_amount", 0)
        total_tax += order.get("total_tax", 0)
        total_base += order.get("total_base", 0)
        for rate_str, info in order.get("tax_breakdown", {}).items():
            if rate_str not in tax_summary:
                tax_summary[rate_str] = {"base": 0, "tax": 0, "count": 0}
            tax_summary[rate_str]["base"] += info.get("base", 0)
            tax_summary[rate_str]["tax"] += info.get("tax", 0)
            tax_summary[rate_str]["count"] += 1
    # Round
    for k in tax_summary:
        tax_summary[k]["base"] = round(tax_summary[k]["base"], 2)
        tax_summary[k]["tax"] = round(tax_summary[k]["tax"], 2)
    return {
        "total_sales": round(total_sales, 2),
        "total_tax": round(total_tax, 2),
        "total_base": round(total_base, 2),
        "breakdown": tax_summary,
        "order_count": len(orders)
    }

# ==================== Report Export (Excel) ====================
@router.get("/reports/export/sales")
async def export_sales_report(start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    import openpyxl
    query = {}
    if start_date: query["created_at"] = {"$gte": start_date}
    if end_date: query.setdefault("created_at", {})["$lte"] = end_date
    orders = await udb.sales_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    headers = ["Order No", "Date", "Customer", "Items", "Total", "Payment", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    customers = {c["id"]: c["name"] for c in await udb.customers.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(10000)}
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.get("order_no", ""))
        ws.cell(row=row, column=2, value=order.get("created_at", "")[:19])
        ws.cell(row=row, column=3, value=customers.get(order.get("customer_id"), "-"))
        ws.cell(row=row, column=4, value=len(order.get("items", [])))
        ws.cell(row=row, column=5, value=order.get("total_amount", 0))
        ws.cell(row=row, column=6, value=order.get("payment_method", ""))
        ws.cell(row=row, column=7, value=order.get("status", ""))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    await log_audit(current_user["user_id"], current_user["username"], "export", "report", "", "Sales report exported", audit_db=udb)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=sales_report.xlsx"})

@router.get("/reports/export/inventory")
async def export_inventory_report(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    import openpyxl
    inventory = await udb.inventory.find({}, {"_id": 0}).to_list(10000)
    products = {p["id"]: p for p in await udb.products.find({}, {"_id": 0}).to_list(10000)}
    warehouses = {w["id"]: w["name"] for w in await udb.warehouses.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    headers = ["Product Code", "Product Name", "Warehouse", "Quantity", "Cost Price", "Total Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    for row, inv in enumerate(inventory, 2):
        product = products.get(inv.get("product_id"), {})
        ws.cell(row=row, column=1, value=product.get("code", ""))
        ws.cell(row=row, column=2, value=product.get("name", ""))
        ws.cell(row=row, column=3, value=warehouses.get(inv.get("warehouse_id"), ""))
        ws.cell(row=row, column=4, value=inv.get("quantity", 0))
        cost = product.get("cost_price", 0)
        ws.cell(row=row, column=5, value=cost)
        ws.cell(row=row, column=6, value=cost * inv.get("quantity", 0))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    await log_audit(current_user["user_id"], current_user["username"], "export", "report", "", "Inventory report exported", audit_db=udb)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=inventory_report.xlsx"})

# ==================== Dashboard Trends ====================
@router.get("/dashboard/trends")
async def get_dashboard_trends(days: int = 7, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    orders = await udb.sales_orders.find({"created_at": {"$gte": since}}, {"_id": 0}).to_list(10000)
    daily = {}
    for order in orders:
        date_key = order.get("created_at", "")[:10]
        if date_key not in daily:
            daily[date_key] = {"date": date_key, "sales": 0, "count": 0, "profit": 0}
        daily[date_key]["sales"] += order.get("total_amount", 0)
        daily[date_key]["count"] += 1
    # Fill missing days
    result = []
    for i in range(days):
        d = (datetime.now(timezone.utc) - timedelta(days=days - 1 - i)).strftime("%Y-%m-%d")
        entry = daily.get(d, {"date": d, "sales": 0, "count": 0, "profit": 0})
        entry["sales"] = round(entry["sales"], 2)
        result.append(entry)
    return result

