from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from pathlib import Path
import os, logging, json, csv, io

from core import db, master_db, client, get_tenant_db, get_user_db, JWT_SECRET, JWT_ALGORITHM, ROOT_DIR
from core.auth import (
    security, hash_password, verify_password, generate_id, generate_order_no,
    create_token, get_current_user, log_audit
)
from core.models import *

router = APIRouter()
UPLOAD_DIR = Path(ROOT_DIR) / "uploads"

# ==================== Product Routes ====================

@router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product_id = generate_id()
    product_data = product.model_dump()
    # Auto-calculate prices from cost_price and margins
    cost = product_data.get("cost_price", 0)
    if cost > 0:
        if product_data.get("margin1", 0) > 0:
            product_data["price1"] = round(cost * (1 + product_data["margin1"] / 100), 2)
        if product_data.get("margin2", 0) > 0:
            product_data["price2"] = round(cost * (1 + product_data["margin2"] / 100), 2)
        if product_data.get("margin3", 0) > 0:
            product_data["price3"] = round(cost * (1 + product_data["margin3"] / 100), 2)
    # Sync compatibility fields
    product_data["retail_price"] = product_data.get("price1", 0)
    product_data["wholesale_price"] = product_data.get("price3", 0)
    product_doc = {
        "id": product_id,
        **product_data,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.products.insert_one(product_doc)
    return ProductResponse(**product_doc)

@router.get("/products", response_model=List[ProductResponse])
async def get_products(
    category_id: Optional[str] = None,
    search: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if category_id:
        query["category_id"] = category_id
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}},
            {"barcode": {"$regex": search, "$options": "i"}}
        ]
    products = await udb.products.find(query, {"_id": 0}).to_list(1000)
    return [ProductResponse(**p) for p in products]

@router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(product_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product = await udb.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    return ProductResponse(**product)

@router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(product_id: str, product: ProductCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product_data = product.model_dump()
    # Auto-calculate prices from cost_price and margins
    cost = product_data.get("cost_price", 0)
    if cost > 0:
        if product_data.get("margin1", 0) > 0:
            product_data["price1"] = round(cost * (1 + product_data["margin1"] / 100), 2)
        if product_data.get("margin2", 0) > 0:
            product_data["price2"] = round(cost * (1 + product_data["margin2"] / 100), 2)
        if product_data.get("margin3", 0) > 0:
            product_data["price3"] = round(cost * (1 + product_data["margin3"] / 100), 2)
    # Sync compatibility fields
    product_data["retail_price"] = product_data.get("price1", 0)
    product_data["wholesale_price"] = product_data.get("price3", 0)
    result = await udb.products.update_one({"id": product_id}, {"$set": product_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="商品不存在")
    # Track cost price changes
    old_product = await udb.products.find_one({"id": product_id}, {"_id": 0})
    if old_product and cost > 0:
        await udb.cost_history.insert_one({
            "id": generate_id(), "product_id": product_id,
            "cost_price": cost, "price1": product_data.get("price1", 0),
            "changed_by": current_user["user_id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    updated = await udb.products.find_one({"id": product_id}, {"_id": 0})
    return ProductResponse(**updated)

@router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="商品不存在")
    return {"message": "删除成功"}

# ==================== Main App Setup ====================

# Product Import
@router.post("/products/import")
async def import_products(file: UploadFile = File(...), mode: str = Query("skip", regex="^(skip|overwrite)$"), current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    content = await file.read()
    filename = file.filename.lower()
    items = []

    try:
        if filename.endswith('.json'):
            items = json.loads(content.decode('utf-8'))
            if isinstance(items, dict):
                items = items.get('products', items.get('data', [items]))
        elif filename.endswith('.csv'):
            text = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            items = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1] if cell.value]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        row_dict[header] = row[idx]
                if any(row_dict.values()):
                    items.append(row_dict)
        else:
            raise HTTPException(400, "Unsupported file format. Use .json, .csv, .xlsx, or .xls")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {str(e)}")

    field_map = {
        '编码': 'code', 'codigo': 'code', 'sku': 'code',
        '名称': 'name', 'nombre': 'name', 'product_name': 'name', '商品名称': 'name',
        '分类': 'category', 'categoria': 'category',
        '单位': 'unit', 'unidad': 'unit',
        '成本价': 'cost_price', 'costo': 'cost_price', 'cost': 'cost_price',
        '利率1': 'margin1', 'margen1': 'margin1',
        '利率2': 'margin2', 'margen2': 'margin2',
        '利率3': 'margin3', 'margen3': 'margin3',
        '价格1': 'price1', 'precio1': 'price1',
        '价格2': 'price2', 'precio2': 'price2',
        '价格3': 'price3', 'precio3': 'price3',
        '每箱数量': 'items_per_box', 'cantidad_caja': 'items_per_box', 'box_qty': 'items_per_box',
        '条码': 'barcode', 'codigo_barra': 'barcode',
        '最低库存': 'min_stock', 'stock_minimo': 'min_stock',
        '批发价': 'wholesale_price', 'precio_mayoreo': 'wholesale_price',
        '状态': 'status', 'estado': 'status',
    }

    created = 0
    updated = 0
    skipped = 0
    failed = 0
    errors = []

    for idx, raw in enumerate(items):
        try:
            item = {}
            for k, v in raw.items():
                if k is None:
                    continue
                mapped = field_map.get(str(k).strip().lower(), str(k).strip().lower())
                item[mapped] = v

            if not item.get('code') and not item.get('name'):
                skipped += 1
                continue

            for num_field in ['cost_price', 'margin1', 'margin2', 'margin3', 'price1', 'price2', 'price3', 'items_per_box', 'min_stock', 'wholesale_price']:
                if num_field in item and item[num_field] is not None:
                    try:
                        item[num_field] = float(str(item[num_field]).replace(',', ''))
                    except (ValueError, TypeError):
                        item[num_field] = 0

            code = item.get('code', '')
            existing = await udb.products.find_one({"code": code}, {"_id": 0}) if code else None

            product_data = {
                "code": item.get('code', ''),
                "name": item.get('name', ''),
                "category": item.get('category', ''),
                "unit": item.get('unit', 'pcs'),
                "cost_price": item.get('cost_price', 0),
                "margin1": item.get('margin1', 30),
                "margin2": item.get('margin2', 20),
                "margin3": item.get('margin3', 10),
                "price1": item.get('price1', 0),
                "price2": item.get('price2', 0),
                "price3": item.get('price3', 0),
                "items_per_box": int(item.get('items_per_box', 1) or 1),
                "barcode": item.get('barcode', ''),
                "min_stock": int(item.get('min_stock', 0) or 0),
                "wholesale_price": item.get('wholesale_price', 0),
                "status": item.get('status', 'active'),
            }

            for m_key in ['margin1', 'margin2', 'margin3']:
                p_key = m_key.replace('margin', 'price')
                if product_data[p_key] == 0 and product_data['cost_price'] > 0 and product_data[m_key] > 0:
                    product_data[p_key] = round(product_data['cost_price'] * (1 + product_data[m_key] / 100), 2)

            if existing:
                if mode == 'overwrite':
                    await udb.products.update_one({"code": code}, {"$set": {**product_data, "updated_at": datetime.now(timezone.utc).isoformat()}})
                    updated += 1
                else:
                    skipped += 1
            else:
                product_data["id"] = str(uuid.uuid4())
                product_data["stock"] = 0
                product_data["created_at"] = datetime.now(timezone.utc).isoformat()
                product_data["updated_at"] = datetime.now(timezone.utc).isoformat()
                await udb.products.insert_one(product_data)
                created += 1
        except Exception as e:
            failed += 1
            errors.append(f"Row {idx + 1}: {str(e)}")

    return {"created": created, "updated": updated, "skipped": skipped, "failed": failed, "errors": errors[:10]}

# Product import template
@router.get("/products/import/template")
async def get_import_template():
    from fastapi.responses import StreamingResponse
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['code', 'name', 'category', 'unit', 'cost_price', 'margin1', 'margin2', 'margin3', 'price1', 'price2', 'price3', 'items_per_box', 'barcode', 'min_stock', 'wholesale_price', 'status'])
    writer.writerow(['P001', 'Sample Product', 'General', 'pcs', '10.00', '30', '20', '10', '13.00', '12.00', '11.00', '12', '7501234567890', '5', '11.50', 'active'])
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=product_import_template.csv"}
    )

@router.get("/")
async def root():
    return {"message": "Sellox API", "status": "running"}

# ==================== 2. Product Image Upload ====================
@router.post("/products/{product_id}/image")
async def upload_product_image(product_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product = await udb.products.find_one({"id": product_id})
    if not product: raise HTTPException(404, "Product not found")
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "jpg"
    if ext not in ("jpg", "jpeg", "png", "webp"): raise HTTPException(400, "Invalid image format")
    filename = f"{product_id}.{ext}"
    filepath = UPLOAD_DIR / "products" / filename
    with open(filepath, "wb") as f:
        content = await file.read()
        f.write(content)
    image_url = f"/api/uploads/products/{filename}"
    await udb.products.update_one({"id": product_id}, {"$set": {"image_url": image_url}})
    await log_audit(current_user["user_id"], current_user["username"], "upload", "product", product_id, "Image uploaded", audit_db=udb)
    return {"image_url": image_url}

@router.delete("/products/{product_id}/image")
async def delete_product_image(product_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product = await udb.products.find_one({"id": product_id}, {"_id": 0})
    if not product: raise HTTPException(404, "Product not found")
    if product.get("image_url"):
        filepath = ROOT_DIR / product["image_url"].lstrip("/")
        if filepath.exists(): filepath.unlink()
    await udb.products.update_one({"id": product_id}, {"$set": {"image_url": ""}})
    return {"message": "Image deleted"}

# ==================== 8. Product Bundles ====================
@router.post("/bundles")
async def create_bundle(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    bundle = {
        "id": generate_id(), "name": data.get("name"),
        "description": data.get("description", ""),
        "items": data.get("items", []),  # [{product_id, quantity}]
        "bundle_price": data.get("bundle_price", 0),
        "original_price": data.get("original_price", 0),
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.bundles.insert_one(bundle)
    del bundle["_id"]
    return bundle

@router.get("/bundles")
async def get_bundles(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    bundles = await udb.bundles.find({}, {"_id": 0}).to_list(200)
    products = {p["id"]: p for p in await udb.products.find({}, {"_id": 0}).to_list(10000)}
    for b in bundles:
        for item in b.get("items", []):
            p = products.get(item.get("product_id"))
            if p: item["product_name"] = p.get("name", "?")
    return bundles

@router.put("/bundles/{bundle_id}")
async def update_bundle(bundle_id: str, data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    await udb.bundles.update_one({"id": bundle_id}, {"$set": data})
    return await udb.bundles.find_one({"id": bundle_id}, {"_id": 0})

@router.delete("/bundles/{bundle_id}")
async def delete_bundle(bundle_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    await udb.bundles.delete_one({"id": bundle_id})
    return {"status": "deleted"}

# ==================== 9. Cost Price Tracking ====================
@router.get("/products/{product_id}/cost-history")
async def get_cost_history(product_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    return await udb.cost_history.find({"product_id": product_id}, {"_id": 0}).sort("created_at", -1).to_list(100)

