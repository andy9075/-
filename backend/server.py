from fastapi import FastAPI, APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from pathlib import Path
import os, logging, json, csv, io, shutil

# Core imports
from core import db, master_db, client, get_tenant_db, get_user_db, JWT_SECRET, JWT_ALGORITHM, ROOT_DIR
from core.auth import (
    security, hash_password, verify_password, generate_id, generate_order_no,
    create_token, get_current_user, log_audit
)
from core.models import (
    UserCreate, UserLogin, UserResponse,
    StoreCreate, StoreResponse,
    WarehouseCreate, WarehouseResponse,
    SupplierCreate, SupplierResponse,
    CustomerCreate, CustomerResponse,
    CategoryCreate, CategoryResponse,
    ProductCreate, ProductResponse,
    InventoryCreate, InventoryResponse, InventoryAdjust,
    PurchaseItemCreate, PurchaseOrderCreate, PurchaseOrderResponse,
    SalesItemCreate, SalesOrderCreate, SalesOrderResponse,
    PaymentSettingsUpdate, SystemSettings,
    OnlineOrderCreate, OnlineOrderResponse,
)

app = FastAPI(title="Sellox API")
api_router = APIRouter(prefix="/api")


# ==================== Auth Routes ====================

@api_router.post("/auth/register", response_model=dict)
async def register(user: UserCreate):
    existing = await db.users.find_one({"username": user.username})
    if existing:
        raise HTTPException(status_code=400, detail="用户名已存在")
    
    user_id = generate_id()
    user_doc = {
        "id": user_id,
        "username": user.username,
        "password": hash_password(user.password),
        "role": user.role,
        "store_id": user.store_id,
        "name": user.name,
        "phone": user.phone,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    token = create_token(user_id, user.username, user.role)
    return {"token": token, "user": {k: v for k, v in user_doc.items() if k != "password"}}

@api_router.post("/auth/login", response_model=dict)
async def login(user: UserLogin):
    db_user = await db.users.find_one({"username": user.username})
    if not db_user or not verify_password(user.password, db_user["password"]):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    
    token = create_token(db_user["id"], db_user["username"], db_user["role"])
    return {"token": token, "user": {k: v for k, v in db_user.items() if k not in ["password", "_id"]}}

@api_router.get("/auth/me", response_model=dict)
async def get_me(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    user = await udb.users.find_one({"id": current_user["user_id"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if current_user.get("tenant_id"):
        user["tenant_id"] = current_user["tenant_id"]
    return user

@api_router.get("/auth/cashiers")
async def get_cashiers(current_user: dict = Depends(get_current_user)):
    """List users for POS login selection"""
    udb = get_user_db(current_user)
    users = await udb.users.find({}, {"_id": 0, "password": 0}).to_list(100)
    return [{"id": u["id"], "username": u["username"], "name": u.get("name", ""), "role": u.get("role", "staff")} for u in users]

# ==================== Store Routes ====================

@api_router.post("/stores", response_model=StoreResponse)
async def create_store(store: StoreCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    store_id = generate_id()
    store_doc = {
        "id": store_id,
        **store.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.stores.insert_one(store_doc)
    return StoreResponse(**store_doc)

@api_router.get("/stores", response_model=List[StoreResponse])
async def get_stores(type: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    query = {}
    if type:
        query["type"] = type
    stores = await udb.stores.find(query, {"_id": 0}).to_list(1000)
    return [StoreResponse(**s) for s in stores]

@api_router.get("/stores/{store_id}", response_model=StoreResponse)
async def get_store(store_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    store = await udb.stores.find_one({"id": store_id}, {"_id": 0})
    if not store:
        raise HTTPException(status_code=404, detail="门店不存在")
    return StoreResponse(**store)

@api_router.put("/stores/{store_id}", response_model=StoreResponse)
async def update_store(store_id: str, store: StoreCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.stores.update_one({"id": store_id}, {"$set": store.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="门店不存在")
    updated = await udb.stores.find_one({"id": store_id}, {"_id": 0})
    return StoreResponse(**updated)

@api_router.delete("/stores/{store_id}")
async def delete_store(store_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.stores.delete_one({"id": store_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="门店不存在")
    return {"message": "删除成功"}

# ==================== Warehouse Routes ====================

@api_router.post("/warehouses", response_model=WarehouseResponse)
async def create_warehouse(warehouse: WarehouseCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    warehouse_id = generate_id()
    warehouse_doc = {
        "id": warehouse_id,
        **warehouse.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.warehouses.insert_one(warehouse_doc)
    return WarehouseResponse(**warehouse_doc)

@api_router.get("/warehouses", response_model=List[WarehouseResponse])
async def get_warehouses(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    warehouses = await udb.warehouses.find({}, {"_id": 0}).to_list(1000)
    return [WarehouseResponse(**w) for w in warehouses]

@api_router.get("/warehouses/{warehouse_id}", response_model=WarehouseResponse)
async def get_warehouse(warehouse_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    warehouse = await udb.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    if not warehouse:
        raise HTTPException(status_code=404, detail="仓库不存在")
    return WarehouseResponse(**warehouse)

@api_router.get("/warehouses/main/info", response_model=WarehouseResponse)
async def get_main_warehouse(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    warehouse = await udb.warehouses.find_one({"is_main": True}, {"_id": 0})
    if not warehouse:
        raise HTTPException(status_code=404, detail="总部仓库不存在")
    return WarehouseResponse(**warehouse)

# ==================== Supplier Routes ====================

@api_router.post("/suppliers", response_model=SupplierResponse)
async def create_supplier(supplier: SupplierCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    supplier_id = generate_id()
    supplier_doc = {
        "id": supplier_id,
        **supplier.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.suppliers.insert_one(supplier_doc)
    return SupplierResponse(**supplier_doc)

@api_router.get("/suppliers", response_model=List[SupplierResponse])
async def get_suppliers(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    suppliers = await udb.suppliers.find({}, {"_id": 0}).to_list(1000)
    return [SupplierResponse(**s) for s in suppliers]

@api_router.put("/suppliers/{supplier_id}", response_model=SupplierResponse)
async def update_supplier(supplier_id: str, supplier: SupplierCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.suppliers.update_one({"id": supplier_id}, {"$set": supplier.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="供应商不存在")
    updated = await udb.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    return SupplierResponse(**updated)

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="供应商不存在")
    return {"message": "删除成功"}

# ==================== Customer Routes ====================

@api_router.post("/customers", response_model=CustomerResponse)
async def create_customer(customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    customer_id = generate_id()
    customer_doc = {
        "id": customer_id,
        **customer.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.customers.insert_one(customer_doc)
    return CustomerResponse(**customer_doc)

@api_router.get("/customers", response_model=List[CustomerResponse])
async def get_customers(search: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}}
        ]
    customers = await udb.customers.find(query, {"_id": 0}).to_list(1000)
    return [CustomerResponse(**c) for c in customers]

@api_router.get("/customers/{customer_id}", response_model=CustomerResponse)
async def get_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    customer = await udb.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    return CustomerResponse(**customer)

@api_router.put("/customers/{customer_id}", response_model=CustomerResponse)
async def update_customer(customer_id: str, customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.customers.update_one({"id": customer_id}, {"$set": customer.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="客户不存在")
    updated = await udb.customers.find_one({"id": customer_id}, {"_id": 0})
    return CustomerResponse(**updated)

# ==================== Category Routes ====================

@api_router.post("/categories", response_model=CategoryResponse)
async def create_category(category: CategoryCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    category_id = generate_id()
    category_doc = {
        "id": category_id,
        **category.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.categories.insert_one(category_doc)
    return CategoryResponse(**category_doc)

@api_router.get("/categories", response_model=List[CategoryResponse])
async def get_categories(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    categories = await udb.categories.find({}, {"_id": 0}).sort("sort_order", 1).to_list(1000)
    return [CategoryResponse(**c) for c in categories]

@api_router.put("/categories/{category_id}", response_model=CategoryResponse)
async def update_category(category_id: str, category: CategoryCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.categories.update_one({"id": category_id}, {"$set": category.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="分类不存在")
    updated = await udb.categories.find_one({"id": category_id}, {"_id": 0})
    return CategoryResponse(**updated)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="分类不存在")
    return {"message": "删除成功"}

# ==================== Product Routes ====================

@api_router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product_id = generate_id()
    product_data = product.model_dump()
    # Auto-calculate prices from cost_price and margins
    cost = product_data.get("cost_price", 0)
    if cost > 0:
        if product_data.get("margin1", 0) > 0:
            product_data["price1"] = round(cost * (1 + product_data["margin1"] / 100), 2)
        if product_data.get("margin2", 0) > 0:
            product_data["price2"] = round(cost * (1 + product_data["margin2"] / 100), 2)
        if product_data.get("margin3", 0) > 0:
            product_data["price3"] = round(cost * (1 + product_data["margin3"] / 100), 2)
    # Sync compatibility fields
    product_data["retail_price"] = product_data.get("price1", 0)
    product_data["wholesale_price"] = product_data.get("price3", 0)
    product_doc = {
        "id": product_id,
        **product_data,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.products.insert_one(product_doc)
    return ProductResponse(**product_doc)

@api_router.get("/products", response_model=List[ProductResponse])
async def get_products(
    category_id: Optional[str] = None,
    search: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if category_id:
        query["category_id"] = category_id
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}},
            {"barcode": {"$regex": search, "$options": "i"}}
        ]
    products = await udb.products.find(query, {"_id": 0}).to_list(1000)
    return [ProductResponse(**p) for p in products]

@api_router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(product_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product = await udb.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    return ProductResponse(**product)

@api_router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(product_id: str, product: ProductCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product_data = product.model_dump()
    # Auto-calculate prices from cost_price and margins
    cost = product_data.get("cost_price", 0)
    if cost > 0:
        if product_data.get("margin1", 0) > 0:
            product_data["price1"] = round(cost * (1 + product_data["margin1"] / 100), 2)
        if product_data.get("margin2", 0) > 0:
            product_data["price2"] = round(cost * (1 + product_data["margin2"] / 100), 2)
        if product_data.get("margin3", 0) > 0:
            product_data["price3"] = round(cost * (1 + product_data["margin3"] / 100), 2)
    # Sync compatibility fields
    product_data["retail_price"] = product_data.get("price1", 0)
    product_data["wholesale_price"] = product_data.get("price3", 0)
    result = await udb.products.update_one({"id": product_id}, {"$set": product_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="商品不存在")
    # Track cost price changes
    old_product = await udb.products.find_one({"id": product_id}, {"_id": 0})
    if old_product and cost > 0:
        await udb.cost_history.insert_one({
            "id": generate_id(), "product_id": product_id,
            "cost_price": cost, "price1": product_data.get("price1", 0),
            "changed_by": current_user["user_id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    updated = await udb.products.find_one({"id": product_id}, {"_id": 0})
    return ProductResponse(**updated)

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="商品不存在")
    return {"message": "删除成功"}

# ==================== Inventory Routes ====================

@api_router.get("/inventory", response_model=List[dict])
async def get_inventory(
    warehouse_id: Optional[str] = None,
    product_id: Optional[str] = None,
    low_stock: bool = False,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    if product_id:
        query["product_id"] = product_id
    
    inventory = await udb.inventory.find(query, {"_id": 0}).to_list(10000)
    
    result = []
    for inv in inventory:
        product = await udb.products.find_one({"id": inv["product_id"]}, {"_id": 0})
        warehouse = await udb.warehouses.find_one({"id": inv["warehouse_id"]}, {"_id": 0})
        inv["product"] = product
        inv["warehouse"] = warehouse
        inv["available"] = inv["quantity"] - inv.get("reserved", 0)
        
        if low_stock and product:
            if inv["quantity"] > product.get("min_stock", 0):
                continue
        result.append(inv)
    
    return result

@api_router.post("/inventory/adjust")
async def adjust_inventory(adjust: InventoryAdjust, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    existing = await udb.inventory.find_one({
        "product_id": adjust.product_id,
        "warehouse_id": adjust.warehouse_id
    })
    
    if existing:
        new_qty = existing["quantity"] + adjust.quantity
        if new_qty < 0:
            raise HTTPException(status_code=400, detail="库存不足")
        await udb.inventory.update_one(
            {"product_id": adjust.product_id, "warehouse_id": adjust.warehouse_id},
            {"$set": {"quantity": new_qty, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        if adjust.quantity < 0:
            raise HTTPException(status_code=400, detail="库存不足")
        await udb.inventory.insert_one({
            "id": generate_id(),
            "product_id": adjust.product_id,
            "warehouse_id": adjust.warehouse_id,
            "quantity": adjust.quantity,
            "reserved": 0,
            "updated_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Record adjustment log
    await udb.inventory_logs.insert_one({
        "id": generate_id(),
        "product_id": adjust.product_id,
        "warehouse_id": adjust.warehouse_id,
        "quantity": adjust.quantity,
        "reason": adjust.reason,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "库存调整成功"}

@api_router.post("/inventory/transfer")
async def transfer_inventory(
    product_id: str,
    from_warehouse_id: str,
    to_warehouse_id: str,
    quantity: int,
    current_user: dict = Depends(get_current_user)
):
    # Check source inventory
    udb = get_user_db(current_user)
    source = await udb.inventory.find_one({
        "product_id": product_id,
        "warehouse_id": from_warehouse_id
    })
    
    if not source or source["quantity"] < quantity:
        raise HTTPException(status_code=400, detail="源仓库库存不足")
    
    # Decrease source
    await udb.inventory.update_one(
        {"product_id": product_id, "warehouse_id": from_warehouse_id},
        {"$inc": {"quantity": -quantity}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Increase destination
    dest = await udb.inventory.find_one({
        "product_id": product_id,
        "warehouse_id": to_warehouse_id
    })
    
    if dest:
        await udb.inventory.update_one(
            {"product_id": product_id, "warehouse_id": to_warehouse_id},
            {"$inc": {"quantity": quantity}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        await udb.inventory.insert_one({
            "id": generate_id(),
            "product_id": product_id,
            "warehouse_id": to_warehouse_id,
            "quantity": quantity,
            "reserved": 0,
            "updated_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Record transfer log
    await udb.transfer_logs.insert_one({
        "id": generate_id(),
        "product_id": product_id,
        "from_warehouse_id": from_warehouse_id,
        "to_warehouse_id": to_warehouse_id,
        "quantity": quantity,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "库存调拨成功"}


@api_router.get("/transfer-logs")
async def get_transfer_logs(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    logs = await udb.transfer_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return logs


# ==================== Purchase Order Routes ====================

@api_router.post("/purchase-orders", response_model=PurchaseOrderResponse)
async def create_purchase_order(order: PurchaseOrderCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    order_id = generate_id()
    order_no = generate_order_no("PO")
    
    total_amount = sum(item.amount for item in order.items)
    
    order_doc = {
        "id": order_id,
        "order_no": order_no,
        "supplier_id": order.supplier_id,
        "warehouse_id": order.warehouse_id,
        "items": [item.model_dump() for item in order.items],
        "total_amount": total_amount,
        "status": "pending",
        "notes": order.notes,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await udb.purchase_orders.insert_one(order_doc)
    return PurchaseOrderResponse(**order_doc)

@api_router.get("/purchase-orders", response_model=List[PurchaseOrderResponse])
async def get_purchase_orders(
    status: Optional[str] = None,
    supplier_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if status:
        query["status"] = status
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    orders = await udb.purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [PurchaseOrderResponse(**o) for o in orders]

@api_router.put("/purchase-orders/{order_id}/receive")
async def receive_purchase_order(order_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    order = await udb.purchase_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="采购单不存在")
    
    if order["status"] != "pending":
        raise HTTPException(status_code=400, detail="订单状态不允许入库")
    
    # Update inventory
    for item in order["items"]:
        existing = await udb.inventory.find_one({
            "product_id": item["product_id"],
            "warehouse_id": order["warehouse_id"]
        })
        
        if existing:
            await udb.inventory.update_one(
                {"product_id": item["product_id"], "warehouse_id": order["warehouse_id"]},
                {"$inc": {"quantity": item["quantity"]}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
            )
        else:
            await udb.inventory.insert_one({
                "id": generate_id(),
                "product_id": item["product_id"],
                "warehouse_id": order["warehouse_id"],
                "quantity": item["quantity"],
                "reserved": 0,
                "updated_at": datetime.now(timezone.utc).isoformat()
            })
    
    # Update order status
    await udb.purchase_orders.update_one(
        {"id": order_id},
        {"$set": {"status": "received", "received_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "入库成功"}

# ==================== Sales Order Routes ====================

@api_router.post("/sales-orders", response_model=SalesOrderResponse)
async def create_sales_order(order: SalesOrderCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    order_id = generate_id()
    order_no = generate_order_no("SO")
    
    total_amount = sum(item.amount for item in order.items)
    discount_amount = sum(item.discount for item in order.items)
    
    # Points redemption: 100 points = $1 (configurable via system settings)
    sys_settings = await udb.system_settings.find_one({"key": "system"}, {"_id": 0}) or {}
    points_per_dollar = sys_settings.get("points_per_dollar", 1)  # earn rate
    points_value_rate = sys_settings.get("points_value_rate", 100)  # 100 points = $1
    points_discount = 0.0
    if order.points_used > 0 and order.customer_id:
        customer = await udb.customers.find_one({"id": order.customer_id}, {"_id": 0})
        if not customer or customer.get("points", 0) < order.points_used:
            raise HTTPException(400, "Insufficient points")
        points_discount = order.points_used / points_value_rate
        if points_discount > total_amount:
            points_discount = total_amount
    
    final_total = total_amount - points_discount
    
    # Get store's warehouse
    store = await udb.stores.find_one({"id": order.store_id})
    warehouse_id = store.get("warehouse_id") if store else None
    
    # Check and deduct inventory
    if warehouse_id:
        for item in order.items:
            inv = await udb.inventory.find_one({
                "product_id": item.product_id,
                "warehouse_id": warehouse_id
            })
            if not inv or inv["quantity"] < item.quantity:
                product = await udb.products.find_one({"id": item.product_id})
                raise HTTPException(status_code=400, detail=f"商品 {product['name'] if product else item.product_id} 库存不足")
            
            await udb.inventory.update_one(
                {"product_id": item.product_id, "warehouse_id": warehouse_id},
                {"$inc": {"quantity": -item.quantity}}
            )
    
    # Build items with product names
    order_items = []
    for item in order.items:
        item_dict = item.model_dump()
        product = await udb.products.find_one({"id": item.product_id})
        if product:
            item_dict["product_name"] = product["name"]
        order_items.append(item_dict)
    
    order_doc = {
        "id": order_id,
        "order_no": order_no,
        "store_id": order.store_id,
        "customer_id": order.customer_id,
        "items": order_items,
        "total_amount": final_total,
        "original_amount": total_amount,
        "discount_amount": discount_amount,
        "points_used": order.points_used,
        "points_discount": round(points_discount, 2),
        "paid_amount": order.paid_amount,
        "payment_method": order.payment_method,
        "status": "completed" if order.paid_amount >= final_total else "pending",
        "notes": order.notes,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await udb.sales_orders.insert_one(order_doc)
    
    # Update customer points if applicable
    points_earned = 0
    if order.customer_id:
        points_earned = int(final_total * points_per_dollar)
        net_points = points_earned - order.points_used
        if net_points != 0:
            await udb.customers.update_one(
                {"id": order.customer_id},
                {"$inc": {"points": net_points}}
            )
        # Log points earned
        if points_earned > 0:
            await udb.points_log.insert_one({
                "id": generate_id(), "customer_id": order.customer_id,
                "type": "earn", "amount": points_earned,
                "reason": f"Purchase {order_no}",
                "created_by": current_user["user_id"],
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        # Log points redeemed
        if order.points_used > 0:
            await udb.points_log.insert_one({
                "id": generate_id(), "customer_id": order.customer_id,
                "type": "redeem", "amount": -order.points_used,
                "reason": f"Redeem at {order_no} (-${points_discount:.2f})",
                "created_by": current_user["user_id"],
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    
    order_doc.pop("_id", None)
    order_doc["points_earned"] = points_earned
    
    # Auto-upgrade customer VIP level
    if order.customer_id:
        rules_doc = await udb.system_settings.find_one({"key": "vip_rules"}, {"_id": 0})
        vip_levels = (rules_doc or {}).get("levels", [
            {"name": "normal", "min_spent": 0}, {"name": "silver", "min_spent": 200},
            {"name": "gold", "min_spent": 500}, {"name": "vip", "min_spent": 1000}
        ])
        vip_levels.sort(key=lambda x: x["min_spent"], reverse=True)
        all_orders = await udb.sales_orders.find({"customer_id": order.customer_id}, {"_id": 0, "total_amount": 1}).to_list(10000)
        total_spent = sum(o.get("total_amount", 0) for o in all_orders)
        for lv in vip_levels:
            if total_spent >= lv["min_spent"]:
                await udb.customers.update_one({"id": order.customer_id}, {"$set": {"member_level": lv["name"]}})
                break
    
    return SalesOrderResponse(**{k: v for k, v in order_doc.items() if k in SalesOrderResponse.model_fields})

@api_router.get("/sales-orders", response_model=List[SalesOrderResponse])
async def get_sales_orders(
    status: Optional[str] = None,
    store_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if status:
        query["status"] = status
    if store_id:
        query["store_id"] = store_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    orders = await udb.sales_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [SalesOrderResponse(**o) for o in orders]


@api_router.get("/sales-report")
async def get_sales_report(
    store_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if store_id:
        query["store_id"] = store_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date + "T23:59:59"
        else:
            query["created_at"] = {"$lte": end_date + "T23:59:59"}
    
    orders = await udb.sales_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    # Enrich with product names and store names
    products_map = {}
    async for p in udb.products.find({}, {"_id": 0, "id": 1, "name": 1, "code": 1}):
        products_map[p["id"]] = p
    
    stores_map = {}
    async for s in udb.stores.find({}, {"_id": 0, "id": 1, "name": 1}):
        stores_map[s["id"]] = s
    
    # Build report
    store_summary = {}
    product_summary = {}
    total_sales = 0
    total_orders = len(orders)
    total_items = 0
    
    for order in orders:
        store_name = stores_map.get(order.get("store_id"), {}).get("name", "Unknown")
        order_total = order.get("total_amount", 0)
        total_sales += order_total
        
        if store_name not in store_summary:
            store_summary[store_name] = {"store_id": order.get("store_id", ""), "name": store_name, "total": 0, "orders": 0, "products": {}}
        store_summary[store_name]["total"] += order_total
        store_summary[store_name]["orders"] += 1
        
        for item in order.get("items", []):
            pid = item.get("product_id", "")
            pname = products_map.get(pid, {}).get("name", item.get("product_name", "Unknown"))
            pcode = products_map.get(pid, {}).get("code", "")
            qty = item.get("quantity", 0)
            amt = item.get("amount", 0)
            total_items += qty
            
            # Per-store product breakdown
            if pid not in store_summary[store_name]["products"]:
                store_summary[store_name]["products"][pid] = {"name": pname, "code": pcode, "quantity": 0, "amount": 0}
            store_summary[store_name]["products"][pid]["quantity"] += qty
            store_summary[store_name]["products"][pid]["amount"] += amt
            
            # Global product summary
            if pid not in product_summary:
                product_summary[pid] = {"name": pname, "code": pcode, "quantity": 0, "amount": 0}
            product_summary[pid]["quantity"] += qty
            product_summary[pid]["amount"] += amt
    
    # Convert to lists
    store_list = []
    for s in store_summary.values():
        s["products"] = list(s["products"].values())
        store_list.append(s)
    
    return {
        "total_sales": total_sales,
        "total_orders": total_orders,
        "total_items": total_items,
        "stores": store_list,
        "products": list(product_summary.values()),
        "orders": [{
            "id": o.get("id"), "order_no": o.get("order_no"), "store_id": o.get("store_id"),
            "store_name": stores_map.get(o.get("store_id"), {}).get("name", ""),
            "total_amount": o.get("total_amount", 0), "payment_method": o.get("payment_method", ""),
            "created_at": o.get("created_at", ""),
            "items": [{
                "product_name": products_map.get(it.get("product_id"), {}).get("name", it.get("product_name", "")),
                "product_code": products_map.get(it.get("product_id"), {}).get("code", ""),
                "quantity": it.get("quantity", 0), "unit_price": it.get("unit_price", 0), "amount": it.get("amount", 0)
            } for it in o.get("items", [])]
        } for o in orders]
    }


# ==================== Online Shop Routes ====================

@api_router.get("/shop/{tenant_id}/products", response_model=List[dict])
async def get_shop_products(
    tenant_id: str,
    category_id: Optional[str] = None,
    search: Optional[str] = None
):
    """Public endpoint for tenant's online shop"""
    tenant = await master_db.tenants.find_one({"id": tenant_id, "status": "active"})
    if not tenant:
        raise HTTPException(404, "Shop not found")
    shop_db = get_tenant_db(tenant_id)
    query = {"status": "active"}
    if category_id:
        query["category_id"] = category_id
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}}
        ]
    products = await shop_db.products.find(query, {"_id": 0}).to_list(1000)
    main_warehouse = await shop_db.warehouses.find_one({"is_main": True})
    result = []
    for product in products:
        inv = None
        if main_warehouse:
            inv = await shop_db.inventory.find_one({"product_id": product["id"], "warehouse_id": main_warehouse["id"]}, {"_id": 0})
        product["stock"] = inv["quantity"] if inv else 0
        result.append(product)
    return result

@api_router.get("/shop/{tenant_id}/categories", response_model=List[CategoryResponse])
async def get_shop_categories(tenant_id: str):
    """Public endpoint for tenant's shop categories"""
    tenant = await master_db.tenants.find_one({"id": tenant_id, "status": "active"})
    if not tenant:
        raise HTTPException(404, "Shop not found")
    shop_db = get_tenant_db(tenant_id)
    categories = await shop_db.categories.find({}, {"_id": 0}).sort("sort_order", 1).to_list(1000)
    return [CategoryResponse(**c) for c in categories]

@api_router.get("/shop/{tenant_id}/info")
async def get_shop_info(tenant_id: str):
    """Public endpoint to get shop name/info"""
    tenant = await master_db.tenants.find_one({"id": tenant_id, "status": "active"}, {"_id": 0})
    if not tenant:
        raise HTTPException(404, "Shop not found")
    return {"id": tenant["id"], "name": tenant.get("name", ""), "contact_phone": tenant.get("contact_phone", "")}

@api_router.post("/shop/{tenant_id}/orders", response_model=OnlineOrderResponse)
async def create_online_order(tenant_id: str, order: OnlineOrderCreate):
    """Create online shop order for a specific tenant"""
    tenant = await master_db.tenants.find_one({"id": tenant_id, "status": "active"})
    if not tenant:
        raise HTTPException(404, "Shop not found")
    shop_db = get_tenant_db(tenant_id)
    order_id = generate_id()
    order_no = generate_order_no("ON")
    main_warehouse = await shop_db.warehouses.find_one({"is_main": True})
    if not main_warehouse:
        raise HTTPException(status_code=400, detail="Warehouse not configured")
    total_amount = sum(item.amount for item in order.items)
    shipping_fee = 0.0 if total_amount >= 100 else 10.0
    for item in order.items:
        inv = await shop_db.inventory.find_one({"product_id": item.product_id, "warehouse_id": main_warehouse["id"]})
        if not inv or inv["quantity"] < item.quantity:
            product = await shop_db.products.find_one({"id": item.product_id})
            raise HTTPException(status_code=400, detail=f"{product['name'] if product else ''} out of stock")
        await shop_db.inventory.update_one({"product_id": item.product_id, "warehouse_id": main_warehouse["id"]}, {"$inc": {"reserved": item.quantity}})
    online_items = []
    for item in order.items:
        item_dict = item.model_dump()
        product = await shop_db.products.find_one({"id": item.product_id})
        if product:
            item_dict["product_name"] = product["name"]
        online_items.append(item_dict)
    order_doc = {
        "id": order_id, "order_no": order_no, "customer_id": order.customer_id,
        "items": online_items, "total_amount": total_amount, "shipping_fee": shipping_fee,
        "shipping_address": order.shipping_address, "shipping_phone": order.shipping_phone,
        "shipping_name": order.shipping_name, "payment_method": order.payment_method,
        "payment_reference": order.payment_reference, "payment_status": "pending",
        "order_status": "pending", "warehouse_id": main_warehouse["id"],
        "notes": order.notes, "created_at": datetime.now(timezone.utc).isoformat()
    }
    await shop_db.online_orders.insert_one(order_doc)
    return OnlineOrderResponse(**order_doc)

@api_router.get("/shop/{tenant_id}/orders/{order_no}")
async def lookup_shop_order(tenant_id: str, order_no: str):
    """Public: look up order status by order number"""
    tenant = await master_db.tenants.find_one({"id": tenant_id, "status": "active"})
    if not tenant:
        raise HTTPException(404, "Shop not found")
    shop_db = get_tenant_db(tenant_id)
    order = await shop_db.online_orders.find_one({"order_no": order_no}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    return order

@api_router.get("/shop/{tenant_id}/exchange-rates")
async def get_shop_exchange_rates(tenant_id: str):
    """Public: get tenant exchange rates for shop"""
    tenant = await master_db.tenants.find_one({"id": tenant_id, "status": "active"})
    if not tenant:
        raise HTTPException(404, "Shop not found")
    shop_db = get_tenant_db(tenant_id)
    settings = await shop_db.settings.find_one({"type": "exchange_rates"}, {"_id": 0})
    if not settings:
        settings = {"type": "exchange_rates", "usd_to_ves": 36.5, "default_currency": "USD", "local_currency": "VES", "local_currency_symbol": "Bs."}
    return settings

@api_router.get("/shop/{tenant_id}/payment-settings")
async def get_shop_payment_settings(tenant_id: str):
    """Public: get tenant payment settings for shop checkout"""
    tenant = await master_db.tenants.find_one({"id": tenant_id, "status": "active"})
    if not tenant:
        raise HTTPException(404, "Shop not found")
    shop_db = get_tenant_db(tenant_id)
    settings = await shop_db.settings.find_one({"type": "payment"}, {"_id": 0})
    if not settings:
        settings = {"type": "payment", "transfer_enabled": True, "transfer_bank_name": "", "transfer_account": "", "mobile_pay_enabled": False}
    return settings

@api_router.get("/shop/orders", response_model=List[OnlineOrderResponse])
async def get_online_orders(
    customer_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    if status:
        query["order_status"] = status
    
    orders = await udb.online_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [OnlineOrderResponse(**o) for o in orders]

@api_router.put("/shop/orders/{order_id}/pay")
async def pay_online_order(order_id: str):
    order = await db.online_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    
    await db.online_orders.update_one(
        {"id": order_id},
        {"$set": {"payment_status": "paid", "order_status": "processing"}}
    )
    
    return {"message": "支付成功"}

@api_router.get("/shop/order-lookup")
async def lookup_order(order_no: Optional[str] = None, phone: Optional[str] = None):
    """Public endpoint - lookup order by order number or phone"""
    if not order_no and not phone:
        raise HTTPException(status_code=400, detail="请提供订单号或手机号")
    
    query = {}
    if order_no:
        query["order_no"] = order_no
    if phone:
        query["shipping_phone"] = phone
    
    orders = await db.online_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Get product details for each order
    result = []
    for order in orders:
        items_with_details = []
        for item in order.get("items", []):
            product = await db.products.find_one({"id": item["product_id"]}, {"_id": 0})
            items_with_details.append({
                **item,
                "product_name": product["name"] if product else "Unknown",
                "product_code": product["code"] if product else ""
            })
        order["items"] = items_with_details
        result.append(order)
    
    return result

@api_router.put("/shop/orders/{order_id}/ship")
async def ship_online_order(order_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    order = await udb.online_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    
    if order["payment_status"] != "paid":
        raise HTTPException(status_code=400, detail="订单未支付")
    
    # Deduct inventory from warehouse
    for item in order["items"]:
        await udb.inventory.update_one(
            {"product_id": item["product_id"], "warehouse_id": order["warehouse_id"]},
            {"$inc": {"quantity": -item["quantity"], "reserved": -item["quantity"]}}
        )
    
    await udb.online_orders.update_one(
        {"id": order_id},
        {"$set": {"order_status": "shipped", "shipped_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "发货成功"}

@api_router.put("/shop/orders/{order_id}/complete")
async def complete_online_order(order_id: str):
    order = await db.online_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    
    await db.online_orders.update_one(
        {"id": order_id},
        {"$set": {"order_status": "completed", "completed_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Update customer points
    if order.get("customer_id"):
        points_earned = int(order["total_amount"] / 10)
        await db.customers.update_one(
            {"id": order["customer_id"]},
            {"$inc": {"points": points_earned}}
        )
    
    return {"message": "订单完成"}

# ==================== Reports Routes ====================

@api_router.get("/reports/sales-summary")
async def get_sales_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    store_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if store_id:
        query["store_id"] = store_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    sales_orders = await udb.sales_orders.find(query, {"_id": 0}).to_list(10000)
    online_orders = await udb.online_orders.find(
        {"order_status": "completed"} if not query else {**query, "order_status": "completed"},
        {"_id": 0}
    ).to_list(10000)
    
    total_sales = sum(o["total_amount"] for o in sales_orders)
    total_online = sum(o["total_amount"] for o in online_orders)
    
    by_method = {}
    for o in sales_orders:
        method = o.get("payment_method", "other")
        if method not in by_method:
            by_method[method] = {"count": 0, "amount": 0.0}
        by_method[method]["count"] += 1
        by_method[method]["amount"] += o.get("total_amount", 0)
    
    return {
        "total_sales": total_sales,
        "total_online_sales": total_online,
        "total_combined": total_sales + total_online,
        "sales_count": len(sales_orders),
        "online_count": len(online_orders),
        "by_payment_method": by_method
    }

@api_router.get("/reports/inventory-summary")
async def get_inventory_summary(
    warehouse_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if warehouse_id:
        query["warehouse_id"] = warehouse_id
    
    inventory = await udb.inventory.find(query, {"_id": 0}).to_list(10000)
    
    total_items = len(inventory)
    total_quantity = sum(i["quantity"] for i in inventory)
    total_value = 0.0
    low_stock_items = []
    
    for inv in inventory:
        product = await udb.products.find_one({"id": inv["product_id"]}, {"_id": 0})
        if product:
            total_value += inv["quantity"] * product.get("cost_price", 0)
            if inv["quantity"] <= product.get("min_stock", 0):
                low_stock_items.append({
                    "product": product,
                    "quantity": inv["quantity"],
                    "min_stock": product.get("min_stock", 0)
                })
    
    return {
        "total_items": total_items,
        "total_quantity": total_quantity,
        "total_value": total_value,
        "low_stock_count": len(low_stock_items),
        "low_stock_items": low_stock_items[:10]
    }

@api_router.get("/reports/top-products")
async def get_top_products(
    limit: int = 10,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    sales_orders = await udb.sales_orders.find(query, {"_id": 0}).to_list(10000)
    
    product_sales = {}
    for order in sales_orders:
        for item in order.get("items", []):
            pid = item["product_id"]
            if pid not in product_sales:
                product_sales[pid] = {"quantity": 0, "amount": 0}
            product_sales[pid]["quantity"] += item["quantity"]
            product_sales[pid]["amount"] += item["amount"]
    
    # Sort by amount
    sorted_products = sorted(product_sales.items(), key=lambda x: x[1]["amount"], reverse=True)[:limit]
    
    result = []
    for pid, stats in sorted_products:
        product = await udb.products.find_one({"id": pid}, {"_id": 0})
        if product:
            result.append({
                "product": product,
                "quantity": stats["quantity"],
                "amount": stats["amount"]
            })
    
    return result

# ==================== Exchange Rate Settings ====================

@api_router.get("/exchange-rates")
async def get_exchange_rates(current_user: dict = Depends(get_current_user)):
    """Get system exchange rates"""
    udb = get_user_db(current_user)
    settings = await udb.settings.find_one({"type": "exchange_rates"}, {"_id": 0})
    if not settings:
        settings = {
            "type": "exchange_rates",
            "usd_to_ves": 36.5,  # USD to Bolivares
            "usd_to_cop": 4000,  # USD to Colombian Pesos
            "default_currency": "USD",
            "local_currency": "VES",
            "local_currency_symbol": "Bs."
        }
    return settings

@api_router.put("/exchange-rates")
async def update_exchange_rates(
    usd_to_ves: float,
    usd_to_cop: float = 4000,
    default_currency: str = "USD",
    local_currency: str = "VES",
    local_currency_symbol: str = "Bs.",
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    """Update system exchange rates"""
    settings_doc = {
        "type": "exchange_rates",
        "usd_to_ves": usd_to_ves,
        "usd_to_cop": usd_to_cop,
        "default_currency": default_currency,
        "local_currency": local_currency,
        "local_currency_symbol": local_currency_symbol,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.settings.update_one(
        {"type": "exchange_rates"},
        {"$set": settings_doc},
        upsert=True
    )
    return {"message": "汇率已更新"}

# ==================== Payment Settings Routes ====================

@api_router.get("/payment-settings")
async def get_payment_settings(current_user: dict = Depends(get_current_user)):
    """Get payment settings"""
    udb = get_user_db(current_user)
    settings = await udb.settings.find_one({"type": "payment"}, {"_id": 0})
    if not settings:
        # Default settings for Venezuela
        settings = {
            "type": "payment",
            "transfer_enabled": True,
            "transfer_bank_name": "Banco de Venezuela",
            "transfer_account_number": "",
            "transfer_account_holder": "",
            "transfer_rif": "",
            "pago_movil_enabled": True,
            "pago_movil_phone": "",
            "pago_movil_bank_code": "0102",
            "pago_movil_cedula": "",
            "whatsapp_number": ""
        }
    return settings

@api_router.put("/payment-settings")
async def update_payment_settings(settings: PaymentSettingsUpdate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    """Update payment settings - admin only"""
    settings_doc = {
        "type": "payment",
        **settings.model_dump(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.settings.update_one(
        {"type": "payment"},
        {"$set": settings_doc},
        upsert=True
    )
    return {"message": "支付设置已更新"}

@api_router.put("/shop/orders/{order_id}/confirm-payment")
async def confirm_order_payment(order_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    """Admin confirms payment received"""
    order = await udb.online_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    
    await udb.online_orders.update_one(
        {"id": order_id},
        {"$set": {
            "payment_status": "paid",
            "order_status": "processing",
            "payment_confirmed_at": datetime.now(timezone.utc).isoformat(),
            "payment_confirmed_by": current_user["user_id"]
        }}
    )
    
    return {"message": "支付已确认"}

# ==================== Dashboard Routes ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    
    # Today's sales
    today_sales = await udb.sales_orders.find({"created_at": {"$gte": today}}, {"_id": 0}).to_list(1000)
    today_online = await udb.online_orders.find({"created_at": {"$gte": today}}, {"_id": 0}).to_list(1000)
    
    # Counts
    products_count = await udb.products.count_documents({})
    customers_count = await udb.customers.count_documents({})
    stores_count = await udb.stores.count_documents({})
    pending_orders = await udb.online_orders.count_documents({"order_status": "pending"})
    
    return {
        "today_sales_amount": sum(o["total_amount"] for o in today_sales),
        "today_sales_count": len(today_sales),
        "today_online_amount": sum(o["total_amount"] for o in today_online),
        "today_online_count": len(today_online),
        "products_count": products_count,
        "customers_count": customers_count,
        "stores_count": stores_count,
        "pending_online_orders": pending_orders
    }

# ==================== System Settings ====================

@api_router.get("/settings/system")
async def get_system_settings(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    settings = await udb.system_settings.find_one({"key": "system"}, {"_id": 0})
    if not settings:
        return SystemSettings().model_dump()
    return {k: v for k, v in settings.items() if k != "key"}

@api_router.put("/settings/system")
async def update_system_settings(settings: SystemSettings, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    data = settings.model_dump()
    data["key"] = "system"
    await udb.system_settings.update_one({"key": "system"}, {"$set": data}, upsert=True)
    return {"message": "Settings updated"}

# ==================== Employee Management ====================

@api_router.get("/employees", response_model=List[UserResponse])
async def get_employees(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    users = await udb.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return users

@api_router.post("/employees")
async def create_employee(user: UserCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    existing = await udb.users.find_one({"username": user.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username exists")
    user_id = generate_id()
    user_doc = {
        "id": user_id,
        "username": user.username,
        "password": hash_password(user.password),
        "role": user.role,
        "store_id": user.store_id,
        "name": user.name,
        "phone": user.phone,
        "permissions": user.permissions,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.users.insert_one(user_doc)
    return {k: v for k, v in user_doc.items() if k not in ("password", "_id")}

@api_router.put("/employees/{user_id}")
async def update_employee(user_id: str, user: UserCreate, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    update_data = {
        "name": user.name, "phone": user.phone, "role": user.role,
        "store_id": user.store_id, "permissions": user.permissions
    }
    if user.password:
        update_data["password"] = hash_password(user.password)
    await udb.users.update_one({"id": user_id}, {"$set": update_data})
    return {"message": "Employee updated"}

@api_router.delete("/employees/{user_id}")
async def delete_employee(user_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    if current_user.get("user_id", current_user.get("id")) == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    await udb.users.delete_one({"id": user_id})
    return {"message": "Employee deleted"}

# ==================== Refund/Return ====================

@api_router.post("/refunds")
async def create_refund(order_no: str = "", items: List[Dict] = [], reason: str = "", current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    order = await udb.sales_orders.find_one({"order_no": order_no}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    refund_id = generate_id()
    refund_amount = sum(item.get("amount", 0) for item in items)
    
    refund_doc = {
        "id": refund_id,
        "refund_no": generate_order_no("RF"),
        "original_order_no": order_no,
        "original_order_id": order["id"],
        "store_id": order.get("store_id"),
        "items": items,
        "refund_amount": refund_amount,
        "reason": reason,
        "status": "completed",
        "created_by": current_user.get("user_id", current_user.get("id")),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.refunds.insert_one(refund_doc)
    
    # Restore inventory
    for item in items:
        warehouse = await udb.warehouses.find_one({"store_id": order.get("store_id")})
        if warehouse:
            await udb.inventory.update_one(
                {"product_id": item["product_id"], "warehouse_id": warehouse["id"]},
                {"$inc": {"quantity": item.get("quantity", 0)}}
            )
    
    return {k: v for k, v in refund_doc.items() if k != "_id"}

@api_router.get("/refunds")
async def get_refunds(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    refunds = await udb.refunds.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return refunds

# ==================== Stock Alerts ====================

@api_router.get("/stock-alerts")
async def get_stock_alerts(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    products = await udb.products.find({"status": "active"}, {"_id": 0}).to_list(5000)
    alerts = []
    for p in products:
        inv_docs = await udb.inventory.find({"product_id": p["id"]}, {"_id": 0}).to_list(100)
        total_stock = sum(i.get("quantity", 0) for i in inv_docs)
        min_stock = p.get("min_stock", 0)
        if min_stock > 0 and total_stock <= min_stock:
            alerts.append({
                "product_id": p["id"], "product_code": p["code"], "product_name": p["name"],
                "current_stock": total_stock, "min_stock": min_stock,
                "level": "critical" if total_stock == 0 else "warning"
            })
    return alerts

# ==================== Reports: Top/Bottom Products ====================

@api_router.get("/reports/bestsellers")
async def get_bestsellers(days: int = 30, limit: int = 20, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    orders = await udb.sales_orders.find({"created_at": {"$gte": since}}, {"_id": 0}).to_list(10000)
    
    product_sales = {}
    for o in orders:
        for item in o.get("items", []):
            pid = item.get("product_id", "")
            if pid not in product_sales:
                product_sales[pid] = {"quantity": 0, "amount": 0, "name": item.get("product_name", pid)}
            product_sales[pid]["quantity"] += item.get("quantity", 0)
            product_sales[pid]["amount"] += item.get("amount", 0)
    
    sorted_products = sorted(product_sales.items(), key=lambda x: x[1]["amount"], reverse=True)
    bestsellers = [{"product_id": k, **v} for k, v in sorted_products[:limit]]
    slowsellers = [{"product_id": k, **v} for k, v in sorted_products[-limit:] if v["amount"] > 0]
    
    return {"bestsellers": bestsellers, "slowsellers": slowsellers}

# ==================== Stock Taking / Inventory Count ====================

@api_router.post("/stock-taking")
async def create_stock_taking(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    taking_id = generate_id()
    taking_doc = {
        "id": taking_id,
        "taking_no": generate_order_no("ST"),
        "warehouse_id": data.get("warehouse_id"),
        "items": data.get("items", []),  # [{product_id, system_qty, actual_qty, difference}]
        "status": data.get("status", "draft"),
        "notes": data.get("notes", ""),
        "created_by": current_user.get("user_id", current_user.get("id")),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.stock_takings.insert_one(taking_doc)
    
    # If confirmed, adjust inventory
    if data.get("status") == "confirmed":
        for item in data.get("items", []):
            diff = item.get("actual_qty", 0) - item.get("system_qty", 0)
            if diff != 0:
                await udb.inventory.update_one(
                    {"product_id": item["product_id"], "warehouse_id": data["warehouse_id"]},
                    {"$inc": {"quantity": diff}}
                )
    
    return {k: v for k, v in taking_doc.items() if k != "_id"}

@api_router.get("/stock-takings")
async def get_stock_takings(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    takings = await udb.stock_takings.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return takings

# ==================== Daily Settlement ====================

@api_router.get("/reports/daily-settlement")
async def daily_settlement(date: Optional[str] = None, store_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    target_date = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    start = f"{target_date}T00:00:00"
    end = f"{target_date}T23:59:59"
    
    query = {"created_at": {"$gte": start, "$lte": end}}
    if store_id:
        query["store_id"] = store_id
    
    sales = await udb.sales_orders.find(query, {"_id": 0}).to_list(10000)
    refunds = await udb.refunds.find({"created_at": {"$gte": start, "$lte": end}}, {"_id": 0}).to_list(1000)
    
    by_method = {}
    for s in sales:
        m = s.get("payment_method", "cash")
        if m not in by_method:
            by_method[m] = {"count": 0, "amount": 0}
        by_method[m]["count"] += 1
        by_method[m]["amount"] += s.get("total_amount", 0)
    
    total_sales = sum(s.get("total_amount", 0) for s in sales)
    total_cost = 0
    for s in sales:
        for item in s.get("items", []):
            product = await udb.products.find_one({"id": item.get("product_id")})
            if product:
                total_cost += product.get("cost_price", 0) * item.get("quantity", 0)
    
    total_refunds = sum(r.get("refund_amount", 0) for r in refunds)
    
    return {
        "date": target_date,
        "total_sales": total_sales,
        "total_orders": len(sales),
        "total_refunds": total_refunds,
        "refund_count": len(refunds),
        "total_cost": total_cost,
        "gross_profit": total_sales - total_cost - total_refunds,
        "by_payment_method": by_method
    }


# ==================== Init Data ====================

@api_router.post("/init-data")
async def init_data():
    """Initialize default data"""
    # Check if admin exists
    admin = await db.users.find_one({"username": "admin"})
    if not admin:
        await db.users.insert_one({
            "id": generate_id(),
            "username": "admin",
            "password": hash_password("admin123"),
            "role": "admin",
            "store_id": None,
            "name": "系统管理员",
            "phone": "",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Check if main warehouse exists
    main_wh = await db.warehouses.find_one({"is_main": True})
    if not main_wh:
        await db.warehouses.insert_one({
            "id": generate_id(),
            "code": "WH001",
            "name": "总部仓库",
            "address": "总部",
            "is_main": True,
            "store_id": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    return {"message": "初始化完成"}

# ==================== Main App Setup ====================

# Product Import
@api_router.post("/products/import")
async def import_products(file: UploadFile = File(...), mode: str = Query("skip", regex="^(skip|overwrite)$"), current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    content = await file.read()
    filename = file.filename.lower()
    items = []

    try:
        if filename.endswith('.json'):
            items = json.loads(content.decode('utf-8'))
            if isinstance(items, dict):
                items = items.get('products', items.get('data', [items]))
        elif filename.endswith('.csv'):
            text = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            items = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1] if cell.value]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        row_dict[header] = row[idx]
                if any(row_dict.values()):
                    items.append(row_dict)
        else:
            raise HTTPException(400, "Unsupported file format. Use .json, .csv, .xlsx, or .xls")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {str(e)}")

    field_map = {
        '编码': 'code', 'codigo': 'code', 'sku': 'code',
        '名称': 'name', 'nombre': 'name', 'product_name': 'name', '商品名称': 'name',
        '分类': 'category', 'categoria': 'category',
        '单位': 'unit', 'unidad': 'unit',
        '成本价': 'cost_price', 'costo': 'cost_price', 'cost': 'cost_price',
        '利率1': 'margin1', 'margen1': 'margin1',
        '利率2': 'margin2', 'margen2': 'margin2',
        '利率3': 'margin3', 'margen3': 'margin3',
        '价格1': 'price1', 'precio1': 'price1',
        '价格2': 'price2', 'precio2': 'price2',
        '价格3': 'price3', 'precio3': 'price3',
        '每箱数量': 'items_per_box', 'cantidad_caja': 'items_per_box', 'box_qty': 'items_per_box',
        '条码': 'barcode', 'codigo_barra': 'barcode',
        '最低库存': 'min_stock', 'stock_minimo': 'min_stock',
        '批发价': 'wholesale_price', 'precio_mayoreo': 'wholesale_price',
        '状态': 'status', 'estado': 'status',
    }

    created = 0
    updated = 0
    skipped = 0
    failed = 0
    errors = []

    for idx, raw in enumerate(items):
        try:
            item = {}
            for k, v in raw.items():
                if k is None:
                    continue
                mapped = field_map.get(str(k).strip().lower(), str(k).strip().lower())
                item[mapped] = v

            if not item.get('code') and not item.get('name'):
                skipped += 1
                continue

            for num_field in ['cost_price', 'margin1', 'margin2', 'margin3', 'price1', 'price2', 'price3', 'items_per_box', 'min_stock', 'wholesale_price']:
                if num_field in item and item[num_field] is not None:
                    try:
                        item[num_field] = float(str(item[num_field]).replace(',', ''))
                    except (ValueError, TypeError):
                        item[num_field] = 0

            code = item.get('code', '')
            existing = await udb.products.find_one({"code": code}, {"_id": 0}) if code else None

            product_data = {
                "code": item.get('code', ''),
                "name": item.get('name', ''),
                "category": item.get('category', ''),
                "unit": item.get('unit', 'pcs'),
                "cost_price": item.get('cost_price', 0),
                "margin1": item.get('margin1', 30),
                "margin2": item.get('margin2', 20),
                "margin3": item.get('margin3', 10),
                "price1": item.get('price1', 0),
                "price2": item.get('price2', 0),
                "price3": item.get('price3', 0),
                "items_per_box": int(item.get('items_per_box', 1) or 1),
                "barcode": item.get('barcode', ''),
                "min_stock": int(item.get('min_stock', 0) or 0),
                "wholesale_price": item.get('wholesale_price', 0),
                "status": item.get('status', 'active'),
            }

            for m_key in ['margin1', 'margin2', 'margin3']:
                p_key = m_key.replace('margin', 'price')
                if product_data[p_key] == 0 and product_data['cost_price'] > 0 and product_data[m_key] > 0:
                    product_data[p_key] = round(product_data['cost_price'] * (1 + product_data[m_key] / 100), 2)

            if existing:
                if mode == 'overwrite':
                    await udb.products.update_one({"code": code}, {"$set": {**product_data, "updated_at": datetime.now(timezone.utc).isoformat()}})
                    updated += 1
                else:
                    skipped += 1
            else:
                product_data["id"] = str(uuid.uuid4())
                product_data["stock"] = 0
                product_data["created_at"] = datetime.now(timezone.utc).isoformat()
                product_data["updated_at"] = datetime.now(timezone.utc).isoformat()
                await udb.products.insert_one(product_data)
                created += 1
        except Exception as e:
            failed += 1
            errors.append(f"Row {idx + 1}: {str(e)}")

    return {"created": created, "updated": updated, "skipped": skipped, "failed": failed, "errors": errors[:10]}

# Product import template
@api_router.get("/products/import/template")
async def get_import_template():
    from fastapi.responses import StreamingResponse
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['code', 'name', 'category', 'unit', 'cost_price', 'margin1', 'margin2', 'margin3', 'price1', 'price2', 'price3', 'items_per_box', 'barcode', 'min_stock', 'wholesale_price', 'status'])
    writer.writerow(['P001', 'Sample Product', 'General', 'pcs', '10.00', '30', '20', '10', '13.00', '12.00', '11.00', '12', '7501234567890', '5', '11.50', 'active'])
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=product_import_template.csv"}
    )

@api_router.get("/")
async def root():
    return {"message": "Sellox API", "status": "running"}

# ==================== Multi-Tenant Management ====================
@api_router.post("/tenants")
async def create_tenant(data: Dict, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin" or current_user.get("tenant_id"):
        raise HTTPException(403, "Super admin only")
    tenant_id = generate_id()[:8]
    tenant = {
        "id": tenant_id, "name": data.get("name", ""),
        "contact_name": data.get("contact_name", ""),
        "contact_phone": data.get("contact_phone", ""),
        "plan": data.get("plan", "basic"),
        "status": "active", "max_users": data.get("max_users", 5),
        "max_stores": data.get("max_stores", 3),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await master_db.tenants.insert_one(tenant)
    # Create default admin user for tenant
    admin_password = data.get("admin_password", "admin123")
    admin_username = data.get("admin_username", f"admin_{tenant_id}")
    tenant_db = get_tenant_db(tenant_id)
    admin_id = generate_id()
    await tenant_db.users.insert_one({
        "id": admin_id, "username": admin_username,
        "password": hash_password(admin_password), "role": "admin",
        "name": data.get("contact_name", "Admin"), "phone": data.get("contact_phone", ""),
        "store_id": None, "permissions": {}, "created_at": datetime.now(timezone.utc).isoformat()
    })
    # Create default store
    store_id = generate_id()
    await tenant_db.stores.insert_one({
        "id": store_id, "code": "S001", "name": data.get("name", "Main Store"),
        "type": "retail", "address": "", "phone": "", "warehouse_id": None,
        "is_headquarters": True, "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    # Create default warehouse
    wh_id = generate_id()
    await tenant_db.warehouses.insert_one({
        "id": wh_id, "code": "WH001", "name": "Main Warehouse",
        "address": "", "is_main": True, "status": "active",
        "store_id": None, "created_at": datetime.now(timezone.utc).isoformat()
    })
    del tenant["_id"]
    return {**tenant, "admin_username": admin_username, "admin_password": admin_password}

@api_router.get("/tenants")
async def get_tenants(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin" or current_user.get("tenant_id"):
        raise HTTPException(403, "Super admin only")
    tenants = await master_db.tenants.find({}, {"_id": 0}).to_list(500)
    for t in tenants:
        tdb = get_tenant_db(t["id"])
        t["users_count"] = await tdb.users.count_documents({})
        t["orders_count"] = await tdb.sales_orders.count_documents({})
    return tenants

@api_router.put("/tenants/{tenant_id}")
async def update_tenant(tenant_id: str, data: Dict, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin" or current_user.get("tenant_id"):
        raise HTTPException(403, "Super admin only")
    data.pop("id", None); data.pop("_id", None)
    result = await master_db.tenants.update_one({"id": tenant_id}, {"$set": data})
    if result.matched_count == 0: raise HTTPException(404, "Tenant not found")
    return await master_db.tenants.find_one({"id": tenant_id}, {"_id": 0})

@api_router.put("/tenants/{tenant_id}/toggle")
async def toggle_tenant(tenant_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin" or current_user.get("tenant_id"):
        raise HTTPException(403, "Super admin only")
    tenant = await master_db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant: raise HTTPException(404, "Tenant not found")
    new_status = "inactive" if tenant.get("status") == "active" else "active"
    await master_db.tenants.update_one({"id": tenant_id}, {"$set": {"status": new_status}})
    return {"status": new_status}

@api_router.post("/auth/tenant-login")
async def tenant_login(data: Dict):
    tenant_id = data.get("tenant_id", "")
    username = data.get("username", "")
    password = data.get("password", "")
    if not tenant_id or not username or not password:
        raise HTTPException(400, "Missing fields")
    tenant = await master_db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant: raise HTTPException(404, "Tenant not found")
    if tenant.get("status") != "active": raise HTTPException(403, "Tenant is inactive")
    tenant_db = get_tenant_db(tenant_id)
    user = await tenant_db.users.find_one({"username": username})
    if not user or not verify_password(password, user["password"]):
        raise HTTPException(401, "Invalid credentials")
    token = create_token(user["id"], user["username"], user.get("role", "staff"), tenant_id)
    return {
        "token": token,
        "user": {"id": user["id"], "username": user["username"], "role": user.get("role"), "name": user.get("name"), "tenant_id": tenant_id, "tenant_name": tenant.get("name")},
        "tenant": {k: v for k, v in tenant.items() if k != "password"}
    }

@api_router.get("/tenants/{tenant_id}/stats")
async def get_tenant_stats(tenant_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin" or current_user.get("tenant_id"):
        raise HTTPException(403, "Super admin only")
    tdb = get_tenant_db(tenant_id)
    return {
        "users": await tdb.users.count_documents({}),
        "products": await tdb.products.count_documents({}),
        "stores": await tdb.stores.count_documents({})
    }

# ==================== Commission Rules & Calculation ====================
@api_router.get("/settings/commission-rules")
async def get_commission_rules(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    rules = await udb.system_settings.find_one({"key": "commission_rules"}, {"_id": 0})
    if not rules:
        return {"tiers": [
            {"name": "base", "min_progress": 0, "rate": 3},
            {"name": "standard", "min_progress": 60, "rate": 5},
            {"name": "excellent", "min_progress": 100, "rate": 8}
        ]}
    return {k: v for k, v in rules.items() if k != "key"}

@api_router.put("/settings/commission-rules")
async def update_commission_rules(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    data["key"] = "commission_rules"
    await udb.system_settings.update_one({"key": "commission_rules"}, {"$set": data}, upsert=True)
    await log_audit(current_user["user_id"], current_user["username"], "update", "commission_rules", "", "Commission rules updated", audit_db=udb)
    return {"message": "Commission rules updated"}

@api_router.get("/reports/commission")
async def get_commission_report(month: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    start = f"{month}-01T00:00:00"
    end_dt = datetime.strptime(f"{month}-01", "%Y-%m-%d")
    if end_dt.month == 12:
        next_month = end_dt.replace(year=end_dt.year + 1, month=1)
    else:
        next_month = end_dt.replace(month=end_dt.month + 1)
    end = next_month.strftime("%Y-%m-%dT00:00:00")
    
    rules_doc = await udb.system_settings.find_one({"key": "commission_rules"}, {"_id": 0})
    tiers = (rules_doc or {}).get("tiers", [
        {"name": "base", "min_progress": 0, "rate": 3},
        {"name": "standard", "min_progress": 60, "rate": 5},
        {"name": "excellent", "min_progress": 100, "rate": 8}
    ])
    tiers.sort(key=lambda x: x["min_progress"], reverse=True)
    
    employees = await udb.users.find({}, {"_id": 0}).to_list(200)
    targets = await udb.sales_targets.find({"start_date": {"$lte": end}, "end_date": {"$gte": start}}, {"_id": 0}).to_list(200)
    target_map = {}
    for t in targets:
        if t.get("employee_id"):
            target_map[t["employee_id"]] = t.get("target_amount", 0)
    
    orders = await udb.sales_orders.find({"created_at": {"$gte": start, "$lt": end}}, {"_id": 0}).to_list(10000)
    sales_by_emp = {}
    for o in orders:
        uid = o.get("created_by", "")
        sales_by_emp[uid] = sales_by_emp.get(uid, 0) + o.get("total_amount", 0)
    
    result = []
    total_commission = 0
    for emp in employees:
        eid = emp["id"]
        sales = round(sales_by_emp.get(eid, 0), 2)
        target = target_map.get(eid, 0)
        progress = round(sales / target * 100, 1) if target > 0 else (100 if sales > 0 else 0)
        rate = 0
        tier_name = ""
        for tier in tiers:
            if progress >= tier["min_progress"]:
                rate = tier["rate"]
                tier_name = tier["name"]
                break
        commission = round(sales * rate / 100, 2)
        total_commission += commission
        result.append({
            "employee_id": eid, "employee_name": emp.get("name") or emp.get("username"),
            "role": emp.get("role", "staff"), "sales": sales, "target": target,
            "progress": progress, "rate": rate, "tier": tier_name,
            "commission": commission
        })
    result.sort(key=lambda x: x["commission"], reverse=True)
    return {"month": month, "employees": result, "total_commission": round(total_commission, 2)}

@api_router.get("/reports/daily-commission")
async def get_daily_commission(date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    start = f"{date}T00:00:00"
    end = f"{date}T23:59:59"
    
    rules_doc = await udb.system_settings.find_one({"key": "commission_rules"}, {"_id": 0})
    tiers = (rules_doc or {}).get("tiers", [
        {"name": "base", "min_progress": 0, "rate": 3},
        {"name": "standard", "min_progress": 60, "rate": 5},
        {"name": "excellent", "min_progress": 100, "rate": 8}
    ])
    base_rate = min(t["rate"] for t in tiers) if tiers else 3
    
    employees = await udb.users.find({}, {"_id": 0}).to_list(200)
    emp_map = {e["id"]: e.get("name") or e.get("username") for e in employees}
    
    orders = await udb.sales_orders.find({"created_at": {"$gte": start, "$lte": end}}, {"_id": 0}).to_list(10000)
    daily = {}
    for o in orders:
        uid = o.get("created_by", "")
        if uid not in daily:
            daily[uid] = {"employee_name": emp_map.get(uid, "?"), "sales": 0, "count": 0}
        daily[uid]["sales"] += o.get("total_amount", 0)
        daily[uid]["count"] += 1
    
    result = []
    for uid, data in daily.items():
        est_commission = round(data["sales"] * base_rate / 100, 2)
        result.append({**data, "employee_id": uid, "sales": round(data["sales"], 2), "estimated_commission": est_commission})
    result.sort(key=lambda x: x["sales"], reverse=True)
    return {"date": date, "employees": result}

# ==================== 1. VIP Auto-Upgrade ====================
@api_router.get("/settings/vip-rules")
async def get_vip_rules(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    rules = await udb.system_settings.find_one({"key": "vip_rules"}, {"_id": 0})
    if not rules:
        return {"levels": [
            {"name": "normal", "label": "普通会员", "min_spent": 0, "points_multiplier": 1, "discount_percent": 0},
            {"name": "silver", "label": "银卡会员", "min_spent": 200, "points_multiplier": 1.5, "discount_percent": 3},
            {"name": "gold", "label": "金卡会员", "min_spent": 500, "points_multiplier": 2, "discount_percent": 5},
            {"name": "vip", "label": "VIP会员", "min_spent": 1000, "points_multiplier": 3, "discount_percent": 8}
        ]}
    return {k: v for k, v in rules.items() if k != "key"}

@api_router.put("/settings/vip-rules")
async def update_vip_rules(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    data["key"] = "vip_rules"
    await udb.system_settings.update_one({"key": "vip_rules"}, {"$set": data}, upsert=True)
    return {"message": "VIP rules updated"}

@api_router.post("/customers/check-upgrades")
async def check_customer_upgrades(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    rules_doc = await udb.system_settings.find_one({"key": "vip_rules"}, {"_id": 0})
    levels = (rules_doc or {}).get("levels", [
        {"name": "normal", "min_spent": 0}, {"name": "silver", "min_spent": 200},
        {"name": "gold", "min_spent": 500}, {"name": "vip", "min_spent": 1000}
    ])
    levels.sort(key=lambda x: x["min_spent"], reverse=True)
    customers = await udb.customers.find({}, {"_id": 0}).to_list(10000)
    upgraded = 0
    for cust in customers:
        orders = await udb.sales_orders.find({"customer_id": cust["id"]}, {"_id": 0, "total_amount": 1}).to_list(10000)
        total_spent = sum(o.get("total_amount", 0) for o in orders)
        new_level = "normal"
        for lv in levels:
            if total_spent >= lv["min_spent"]:
                new_level = lv["name"]
                break
        if new_level != cust.get("member_level", "normal"):
            await udb.customers.update_one({"id": cust["id"]}, {"$set": {"member_level": new_level}})
            upgraded += 1
    return {"upgraded": upgraded, "total": len(customers)}

# ==================== 2. Product Image Upload ====================
@api_router.post("/products/{product_id}/image")
async def upload_product_image(product_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product = await udb.products.find_one({"id": product_id})
    if not product: raise HTTPException(404, "Product not found")
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "jpg"
    if ext not in ("jpg", "jpeg", "png", "webp"): raise HTTPException(400, "Invalid image format")
    filename = f"{product_id}.{ext}"
    filepath = UPLOAD_DIR / "products" / filename
    with open(filepath, "wb") as f:
        content = await file.read()
        f.write(content)
    image_url = f"/uploads/products/{filename}"
    await udb.products.update_one({"id": product_id}, {"$set": {"image_url": image_url}})
    await log_audit(current_user["user_id"], current_user["username"], "upload", "product", product_id, "Image uploaded", audit_db=udb)
    return {"image_url": image_url}

@api_router.delete("/products/{product_id}/image")
async def delete_product_image(product_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    product = await udb.products.find_one({"id": product_id}, {"_id": 0})
    if not product: raise HTTPException(404, "Product not found")
    if product.get("image_url"):
        filepath = ROOT_DIR / product["image_url"].lstrip("/")
        if filepath.exists(): filepath.unlink()
    await udb.products.update_one({"id": product_id}, {"$set": {"image_url": ""}})
    return {"message": "Image deleted"}

# ==================== 3. Wholesale Module ====================
@api_router.post("/wholesale-orders")
async def create_wholesale_order(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    order_id = generate_id()
    order_no = generate_order_no("WS")
    total = sum(item.get("amount", 0) for item in data.get("items", []))
    order_doc = {
        "id": order_id, "order_no": order_no, "customer_id": data.get("customer_id"),
        "items": data.get("items", []), "total_amount": total,
        "payment_method": data.get("payment_method", "cash"),
        "paid_amount": data.get("paid_amount", 0), "status": "completed" if data.get("paid_amount", 0) >= total else "pending",
        "notes": data.get("notes", ""), "type": "wholesale",
        "created_by": current_user["user_id"], "created_at": datetime.now(timezone.utc).isoformat()
    }
    # Deduct inventory from main warehouse
    warehouse = await udb.warehouses.find_one({"is_main": True}, {"_id": 0})
    wh_id = warehouse["id"] if warehouse else None
    if wh_id:
        for item in data.get("items", []):
            await udb.inventory.update_one(
                {"product_id": item["product_id"], "warehouse_id": wh_id},
                {"$inc": {"quantity": -item.get("quantity", 0)}}
            )
    await udb.wholesale_orders.insert_one(order_doc)
    del order_doc["_id"]
    await log_audit(current_user["user_id"], current_user["username"], "create", "wholesale", order_id, f"${total}", audit_db=udb)
    return order_doc

@api_router.get("/wholesale-orders")
async def get_wholesale_orders(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    return await udb.wholesale_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)

# ==================== 4. Data Restore ====================
@api_router.post("/backup/restore")
async def restore_backup(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin": raise HTTPException(403, "Admin only")
    content = await file.read()
    try: backup = json.loads(content)
    except: raise HTTPException(400, "Invalid JSON file")
    collections = ["products", "categories", "customers", "suppliers", "stores", "warehouses", "inventory", "sales_orders", "purchase_orders", "online_orders", "employees", "promotions", "accounts", "exchange_rates", "payment_settings", "system_settings"]
    restored = {}
    for col_name in collections:
        if col_name in backup and isinstance(backup[col_name], list):
            col = db[col_name]
            if len(backup[col_name]) > 0:
                await col.delete_many({})
                for doc in backup[col_name]:
                    doc.pop("_id", None)
                await col.insert_many(backup[col_name])
            restored[col_name] = len(backup[col_name])
    await log_audit(current_user["user_id"], current_user["username"], "restore", "system", "", f"Restored {len(restored)} collections", audit_db=udb)
    return {"message": "Backup restored", "collections": restored}

# ==================== 5. Employee Attendance ====================
@api_router.post("/attendance/clock-in")
async def clock_in(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    existing = await udb.attendance.find_one({"user_id": current_user["user_id"], "date": today})
    if existing: raise HTTPException(400, "Already clocked in today")
    record = {
        "id": generate_id(), "user_id": current_user["user_id"],
        "username": current_user["username"], "date": today,
        "clock_in": datetime.now(timezone.utc).isoformat(), "clock_out": None,
        "hours": 0, "status": "present"
    }
    await udb.attendance.insert_one(record)
    del record["_id"]
    return record

@api_router.post("/attendance/clock-out")
async def clock_out(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    record = await udb.attendance.find_one({"user_id": current_user["user_id"], "date": today})
    if not record: raise HTTPException(400, "Not clocked in")
    if record.get("clock_out"): raise HTTPException(400, "Already clocked out")
    now = datetime.now(timezone.utc)
    clock_in_time = datetime.fromisoformat(record["clock_in"])
    hours = round((now - clock_in_time).total_seconds() / 3600, 2)
    await udb.attendance.update_one({"id": record["id"]}, {"$set": {"clock_out": now.isoformat(), "hours": hours}})
    return {"clock_out": now.isoformat(), "hours": hours}

@api_router.get("/attendance")
async def get_attendance(start_date: Optional[str] = None, end_date: Optional[str] = None, user_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    query = {}
    if user_id: query["user_id"] = user_id
    if start_date: query["date"] = {"$gte": start_date}
    if end_date: query.setdefault("date", {})["$lte"] = end_date
    records = await udb.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return records

# ==================== 6. Sales Targets ====================
@api_router.post("/sales-targets")
async def create_sales_target(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    target = {
        "id": generate_id(), **data,
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.sales_targets.insert_one(target)
    del target["_id"]
    return target

@api_router.get("/sales-targets")
async def get_sales_targets(period: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    query = {}
    if period: query["period"] = period
    targets = await udb.sales_targets.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    # Calculate progress
    for t in targets:
        start = t.get("start_date", "")
        end = t.get("end_date", "")
        q = {}
        if start: q["created_at"] = {"$gte": start}
        if end: q.setdefault("created_at", {})["$lte"] = end + "T23:59:59"
        if t.get("employee_id"):
            q["created_by"] = t["employee_id"]
        if t.get("store_id"):
            q["store_id"] = t["store_id"]
        orders = await udb.sales_orders.find(q, {"_id": 0, "total_amount": 1}).to_list(10000)
        actual = sum(o.get("total_amount", 0) for o in orders)
        t["actual"] = round(actual, 2)
        t["progress"] = round(actual / t.get("target_amount", 1) * 100, 1) if t.get("target_amount") else 0
    return targets

@api_router.delete("/sales-targets/{target_id}")
async def delete_sales_target(target_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    await udb.sales_targets.delete_one({"id": target_id})
    return {"status": "deleted"}

# ==================== 7. Purchase Returns ====================
@api_router.post("/purchase-returns")
async def create_purchase_return(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    return_id = generate_id()
    return_no = generate_order_no("PR")
    total = sum(item.get("amount", 0) for item in data.get("items", []))
    doc = {
        "id": return_id, "return_no": return_no,
        "supplier_id": data.get("supplier_id"), "purchase_order_id": data.get("purchase_order_id"),
        "items": data.get("items", []), "total_amount": total,
        "reason": data.get("reason", ""), "status": "pending",
        "created_by": current_user["user_id"], "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.purchase_returns.insert_one(doc)
    del doc["_id"]
    # Deduct inventory
    warehouse = await udb.warehouses.find_one({"is_main": True}, {"_id": 0})
    if warehouse:
        for item in data.get("items", []):
            await udb.inventory.update_one(
                {"product_id": item["product_id"], "warehouse_id": warehouse["id"]},
                {"$inc": {"quantity": -item.get("quantity", 0)}}
            )
    await log_audit(current_user["user_id"], current_user["username"], "create", "purchase_return", return_id, f"${total}", audit_db=udb)
    return doc

@api_router.get("/purchase-returns")
async def get_purchase_returns(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    return await udb.purchase_returns.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)

@api_router.put("/purchase-returns/{return_id}/approve")
async def approve_purchase_return(return_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.purchase_returns.update_one({"id": return_id}, {"$set": {"status": "approved"}})
    if result.matched_count == 0: raise HTTPException(404, "Return not found")
    return {"status": "approved"}

# ==================== 8. Product Bundles ====================
@api_router.post("/bundles")
async def create_bundle(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    bundle = {
        "id": generate_id(), "name": data.get("name"),
        "description": data.get("description", ""),
        "items": data.get("items", []),  # [{product_id, quantity}]
        "bundle_price": data.get("bundle_price", 0),
        "original_price": data.get("original_price", 0),
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.bundles.insert_one(bundle)
    del bundle["_id"]
    return bundle

@api_router.get("/bundles")
async def get_bundles(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    bundles = await udb.bundles.find({}, {"_id": 0}).to_list(200)
    products = {p["id"]: p for p in await udb.products.find({}, {"_id": 0}).to_list(10000)}
    for b in bundles:
        for item in b.get("items", []):
            p = products.get(item.get("product_id"))
            if p: item["product_name"] = p.get("name", "?")
    return bundles

@api_router.put("/bundles/{bundle_id}")
async def update_bundle(bundle_id: str, data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    await udb.bundles.update_one({"id": bundle_id}, {"$set": data})
    return await udb.bundles.find_one({"id": bundle_id}, {"_id": 0})

@api_router.delete("/bundles/{bundle_id}")
async def delete_bundle(bundle_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    await udb.bundles.delete_one({"id": bundle_id})
    return {"status": "deleted"}

# ==================== 9. Cost Price Tracking ====================
@api_router.get("/products/{product_id}/cost-history")
async def get_cost_history(product_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    return await udb.cost_history.find({"product_id": product_id}, {"_id": 0}).sort("created_at", -1).to_list(100)

# ==================== 10. Notification Center ====================
@api_router.get("/notifications")
async def get_notifications(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    notifications = []
    # Low stock alerts
    inventory = await udb.inventory.find({}, {"_id": 0}).to_list(10000)
    products = {p["id"]: p for p in await udb.products.find({}, {"_id": 0}).to_list(10000)}
    for inv in inventory:
        product = products.get(inv.get("product_id"))
        if product and inv.get("quantity", 0) <= product.get("min_stock", 5):
            notifications.append({"type": "stock_low", "severity": "warning", "title": f"{product['name']} low stock", "detail": f"Current: {inv['quantity']}, Min: {product.get('min_stock', 5)}", "product_id": inv["product_id"]})
    # Overdue accounts
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    overdue = await udb.accounts.find({"status": {"$ne": "paid"}, "due_date": {"$lt": today}}, {"_id": 0}).to_list(100)
    for acc in overdue:
        notifications.append({"type": "overdue_account", "severity": "error", "title": f"Overdue: {acc.get('party_name')}", "detail": f"${acc.get('amount', 0)} due {acc.get('due_date')}", "account_id": acc["id"]})
    # Pending online orders
    pending_online = await udb.online_orders.count_documents({"order_status": {"$in": ["pending", "confirmed"]}})
    if pending_online > 0:
        notifications.append({"type": "pending_orders", "severity": "info", "title": f"{pending_online} pending online orders", "detail": "Orders waiting for processing"})
    return {"count": len(notifications), "items": notifications}

# ==================== Audit Log ====================

@api_router.get("/audit-logs")
async def get_audit_logs(
    action: Optional[str] = None, target_type: Optional[str] = None,
    start_date: Optional[str] = None, end_date: Optional[str] = None,
    page: int = 1, limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if action: query["action"] = action
    if target_type: query["target_type"] = target_type
    if start_date: query["created_at"] = {"$gte": start_date}
    if end_date: query.setdefault("created_at", {})["$lte"] = end_date
    total = await udb.audit_logs.count_documents(query)
    logs = await udb.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    return {"total": total, "page": page, "limit": limit, "items": logs}

# ==================== Profit Analysis ====================
@api_router.get("/reports/profit-analysis")
async def get_profit_analysis(
    start_date: Optional[str] = None, end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    udb = get_user_db(current_user)
    query = {}
    if start_date: query["created_at"] = {"$gte": start_date}
    if end_date: query.setdefault("created_at", {})["$lte"] = end_date
    orders = await udb.sales_orders.find(query, {"_id": 0}).to_list(10000)
    product_profit = {}
    for order in orders:
        for item in order.get("items", []):
            pid = item["product_id"]
            if pid not in product_profit:
                product_profit[pid] = {"revenue": 0, "cost": 0, "quantity": 0}
            product_profit[pid]["revenue"] += item.get("amount", 0)
            product_profit[pid]["quantity"] += item.get("quantity", 0)
    products = await udb.products.find({}, {"_id": 0}).to_list(10000)
    prod_map = {p["id"]: p for p in products}
    result = []
    total_revenue = 0
    total_cost = 0
    for pid, data in product_profit.items():
        product = prod_map.get(pid, {})
        cost = product.get("cost_price", 0) * data["quantity"]
        profit = data["revenue"] - cost
        margin = (profit / data["revenue"] * 100) if data["revenue"] > 0 else 0
        total_revenue += data["revenue"]
        total_cost += cost
        result.append({
            "product_id": pid, "product_name": product.get("name", "?"),
            "product_code": product.get("code", ""),
            "quantity": data["quantity"], "revenue": round(data["revenue"], 2),
            "cost": round(cost, 2), "profit": round(profit, 2), "margin": round(margin, 1)
        })
    result.sort(key=lambda x: x["profit"], reverse=True)
    total_profit = total_revenue - total_cost
    overall_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    return {
        "items": result, "total_revenue": round(total_revenue, 2),
        "total_cost": round(total_cost, 2), "total_profit": round(total_profit, 2),
        "overall_margin": round(overall_margin, 1)
    }

# ==================== Customer Purchase History & Points/Balance ====================
@api_router.get("/customers/{customer_id}/purchase-history")
async def get_customer_purchase_history(customer_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    orders = await udb.sales_orders.find({"customer_id": customer_id}, {"_id": 0}).sort("created_at", -1).to_list(500)
    total_spent = sum(o.get("total_amount", 0) for o in orders)
    return {"orders": orders, "total_spent": round(total_spent, 2), "order_count": len(orders)}

@api_router.post("/customers/{customer_id}/points/add")
async def add_customer_points(customer_id: str, amount: int = Query(..., gt=0), reason: str = "", current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    customer = await udb.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer: raise HTTPException(404, "Customer not found")
    new_points = customer.get("points", 0) + amount
    await udb.customers.update_one({"id": customer_id}, {"$set": {"points": new_points}})
    await udb.points_log.insert_one({"id": generate_id(), "customer_id": customer_id, "type": "earn", "amount": amount, "balance": new_points, "reason": reason, "created_by": current_user["user_id"], "created_at": datetime.now(timezone.utc).isoformat()})
    await log_audit(current_user["user_id"], current_user["username"], "points_add", "customer", customer_id, f"+{amount} points", audit_db=udb)
    return {"points": new_points}

@api_router.post("/customers/{customer_id}/points/redeem")
async def redeem_customer_points(customer_id: str, amount: int = Query(..., gt=0), reason: str = "", current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    customer = await udb.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer: raise HTTPException(404, "Customer not found")
    current_points = customer.get("points", 0)
    if current_points < amount: raise HTTPException(400, "Insufficient points")
    new_points = current_points - amount
    await udb.customers.update_one({"id": customer_id}, {"$set": {"points": new_points}})
    await udb.points_log.insert_one({"id": generate_id(), "customer_id": customer_id, "type": "redeem", "amount": -amount, "balance": new_points, "reason": reason, "created_by": current_user["user_id"], "created_at": datetime.now(timezone.utc).isoformat()})
    await log_audit(current_user["user_id"], current_user["username"], "points_redeem", "customer", customer_id, f"-{amount} points", audit_db=udb)
    return {"points": new_points}

@api_router.post("/customers/{customer_id}/balance/topup")
async def topup_customer_balance(customer_id: str, amount: float = Query(..., gt=0), reason: str = "", current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    customer = await udb.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer: raise HTTPException(404, "Customer not found")
    new_balance = round(customer.get("balance", 0) + amount, 2)
    await udb.customers.update_one({"id": customer_id}, {"$set": {"balance": new_balance}})
    await udb.balance_log.insert_one({"id": generate_id(), "customer_id": customer_id, "type": "topup", "amount": amount, "balance": new_balance, "reason": reason, "created_by": current_user["user_id"], "created_at": datetime.now(timezone.utc).isoformat()})
    await log_audit(current_user["user_id"], current_user["username"], "balance_topup", "customer", customer_id, f"+${amount}", audit_db=udb)
    return {"balance": new_balance}

@api_router.get("/customers/{customer_id}/points-log")
async def get_customer_points_log(customer_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    logs = await udb.points_log.find({"customer_id": customer_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return logs

@api_router.get("/customers/{customer_id}/balance-log")
async def get_customer_balance_log(customer_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    logs = await udb.balance_log.find({"customer_id": customer_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return logs

# ==================== Report Export (Excel) ====================
@api_router.get("/reports/export/sales")
async def export_sales_report(start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    import openpyxl
    query = {}
    if start_date: query["created_at"] = {"$gte": start_date}
    if end_date: query.setdefault("created_at", {})["$lte"] = end_date
    orders = await udb.sales_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    headers = ["Order No", "Date", "Customer", "Items", "Total", "Payment", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    customers = {c["id"]: c["name"] for c in await udb.customers.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(10000)}
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.get("order_no", ""))
        ws.cell(row=row, column=2, value=order.get("created_at", "")[:19])
        ws.cell(row=row, column=3, value=customers.get(order.get("customer_id"), "-"))
        ws.cell(row=row, column=4, value=len(order.get("items", [])))
        ws.cell(row=row, column=5, value=order.get("total_amount", 0))
        ws.cell(row=row, column=6, value=order.get("payment_method", ""))
        ws.cell(row=row, column=7, value=order.get("status", ""))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    await log_audit(current_user["user_id"], current_user["username"], "export", "report", "", "Sales report exported", audit_db=udb)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=sales_report.xlsx"})

@api_router.get("/reports/export/inventory")
async def export_inventory_report(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    import openpyxl
    inventory = await udb.inventory.find({}, {"_id": 0}).to_list(10000)
    products = {p["id"]: p for p in await udb.products.find({}, {"_id": 0}).to_list(10000)}
    warehouses = {w["id"]: w["name"] for w in await udb.warehouses.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    headers = ["Product Code", "Product Name", "Warehouse", "Quantity", "Cost Price", "Total Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    for row, inv in enumerate(inventory, 2):
        product = products.get(inv.get("product_id"), {})
        ws.cell(row=row, column=1, value=product.get("code", ""))
        ws.cell(row=row, column=2, value=product.get("name", ""))
        ws.cell(row=row, column=3, value=warehouses.get(inv.get("warehouse_id"), ""))
        ws.cell(row=row, column=4, value=inv.get("quantity", 0))
        cost = product.get("cost_price", 0)
        ws.cell(row=row, column=5, value=cost)
        ws.cell(row=row, column=6, value=cost * inv.get("quantity", 0))
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    await log_audit(current_user["user_id"], current_user["username"], "export", "report", "", "Inventory report exported", audit_db=udb)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=inventory_report.xlsx"})

# ==================== Promotions ====================
@api_router.post("/promotions")
async def create_promotion(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    promo = {
        "id": generate_id(), **data,
        "status": data.get("status", "active"),
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.promotions.insert_one(promo)
    del promo["_id"]
    await log_audit(current_user["user_id"], current_user["username"], "create", "promotion", promo["id"], promo.get("name", ""), audit_db=udb)
    return promo

@api_router.get("/promotions")
async def get_promotions(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    query = {}
    if status: query["status"] = status
    return await udb.promotions.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)

@api_router.put("/promotions/{promo_id}")
async def update_promotion(promo_id: str, data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.promotions.update_one({"id": promo_id}, {"$set": data})
    if result.matched_count == 0: raise HTTPException(404, "Promotion not found")
    await log_audit(current_user["user_id"], current_user["username"], "update", "promotion", promo_id, "", audit_db=udb)
    return await udb.promotions.find_one({"id": promo_id}, {"_id": 0})

@api_router.delete("/promotions/{promo_id}")
async def delete_promotion(promo_id: str, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    result = await udb.promotions.delete_one({"id": promo_id})
    if result.deleted_count == 0: raise HTTPException(404, "Promotion not found")
    await log_audit(current_user["user_id"], current_user["username"], "delete", "promotion", promo_id, "", audit_db=udb)
    return {"status": "deleted"}

# ==================== Accounts Receivable / Payable ====================
@api_router.post("/accounts/receivable")
async def create_receivable(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    record = {
        "id": generate_id(), **data, "type": "receivable",
        "status": data.get("status", "pending"),
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.accounts.insert_one(record)
    del record["_id"]
    await log_audit(current_user["user_id"], current_user["username"], "create", "receivable", record["id"], f"${data.get('amount', 0)}", audit_db=udb)
    return record

@api_router.get("/accounts/receivable")
async def get_receivables(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    query = {"type": "receivable"}
    if status: query["status"] = status
    return await udb.accounts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)

@api_router.put("/accounts/{account_id}/pay")
async def pay_account(account_id: str, amount: float = Query(..., gt=0), current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    account = await udb.accounts.find_one({"id": account_id}, {"_id": 0})
    if not account: raise HTTPException(404, "Account not found")
    paid = account.get("paid_amount", 0) + amount
    status = "paid" if paid >= account.get("amount", 0) else "partial"
    await udb.accounts.update_one({"id": account_id}, {"$set": {"paid_amount": paid, "status": status}})
    await log_audit(current_user["user_id"], current_user["username"], "payment", "account", account_id, f"${amount}", audit_db=udb)
    return {"paid_amount": paid, "status": status}

@api_router.post("/accounts/payable")
async def create_payable(data: Dict, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    record = {
        "id": generate_id(), **data, "type": "payable",
        "status": data.get("status", "pending"),
        "created_by": current_user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await udb.accounts.insert_one(record)
    del record["_id"]
    await log_audit(current_user["user_id"], current_user["username"], "create", "payable", record["id"], f"${data.get('amount', 0)}", audit_db=udb)
    return record

@api_router.get("/accounts/payable")
async def get_payables(status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    query = {"type": "payable"}
    if status: query["status"] = status
    return await udb.accounts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)

# ==================== Data Backup ====================
@api_router.get("/backup/export")
async def export_backup(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    collections = ["products", "categories", "customers", "suppliers", "stores", "warehouses", "inventory", "sales_orders", "purchase_orders", "online_orders", "employees", "promotions", "accounts", "exchange_rates", "payment_settings", "system_settings"]
    backup = {}
    for col_name in collections:
        col = db[col_name]
        docs = await col.find({}, {"_id": 0}).to_list(50000)
        backup[col_name] = docs
    backup["exported_at"] = datetime.now(timezone.utc).isoformat()
    backup["version"] = "1.0"
    output = io.BytesIO(json.dumps(backup, ensure_ascii=False, indent=2, default=str).encode("utf-8"))
    await log_audit(current_user["user_id"], current_user["username"], "backup", "system", "", "Full database backup", audit_db=udb)
    return StreamingResponse(output, media_type="application/json", headers={"Content-Disposition": f"attachment; filename=pos_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"})

# ==================== Dashboard Trends ====================
@api_router.get("/dashboard/trends")
async def get_dashboard_trends(days: int = 7, current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    orders = await udb.sales_orders.find({"created_at": {"$gte": since}}, {"_id": 0}).to_list(10000)
    daily = {}
    for order in orders:
        date_key = order.get("created_at", "")[:10]
        if date_key not in daily:
            daily[date_key] = {"date": date_key, "sales": 0, "count": 0, "profit": 0}
        daily[date_key]["sales"] += order.get("total_amount", 0)
        daily[date_key]["count"] += 1
    # Fill missing days
    result = []
    for i in range(days):
        d = (datetime.now(timezone.utc) - timedelta(days=days - 1 - i)).strftime("%Y-%m-%d")
        entry = daily.get(d, {"date": d, "sales": 0, "count": 0, "profit": 0})
        entry["sales"] = round(entry["sales"], 2)
        result.append(entry)
    return result

# ==================== Role Permission Check ====================
@api_router.get("/auth/permissions")
async def get_permissions(current_user: dict = Depends(get_current_user)):
    udb = get_user_db(current_user)
    user = await udb.users.find_one({"id": current_user["user_id"]}, {"_id": 0})
    if not user: raise HTTPException(404, "User not found")
    role = user.get("role", "staff")
    perms = user.get("permissions", {})
    role_defaults = {
        "admin": {"can_discount": True, "can_refund": True, "can_void": True, "can_export": True, "can_manage_employees": True, "can_view_reports": True, "can_manage_settings": True, "max_discount": 100},
        "manager": {"can_discount": True, "can_refund": True, "can_void": True, "can_export": True, "can_manage_employees": False, "can_view_reports": True, "can_manage_settings": False, "max_discount": 50},
        "cashier": {"can_discount": False, "can_refund": False, "can_void": False, "can_export": False, "can_manage_employees": False, "can_view_reports": False, "can_manage_settings": False, "max_discount": 0},
        "staff": {"can_discount": False, "can_refund": False, "can_void": False, "can_export": False, "can_manage_employees": False, "can_view_reports": False, "can_manage_settings": False, "max_discount": 0},
    }
    defaults = role_defaults.get(role, role_defaults["staff"])
    defaults.update(perms)
    return {"role": role, "permissions": defaults}

app.include_router(api_router)

# Serve uploaded files
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
(UPLOAD_DIR / "products").mkdir(exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
